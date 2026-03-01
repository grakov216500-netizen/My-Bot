# server.py — FastAPI сервер для Mini App (финальная версия, с исправлением группы и опросником)

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, Response
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from datetime import datetime, timedelta
import os
import random
import sqlite3
import statistics  # для расчёта медианы
import tempfile
import threading
import time
import urllib.request
import json

from db import get_db, execute, DBIntegrityError

# Импортируем функцию расчёта курса
from utils.course_calculator import get_current_course
# apex_parser импортируем лениво в _get_apex_parser(), чтобы сервер стартовал даже без APEX_USER/PASS

app = FastAPI()

# === CORS: Mini App на GitHub Pages и локальная разработка ===
# При allow_credentials=True нельзя использовать "*" — указываем явные origins
CORS_ORIGINS = [
    "https://grakov216500-netizen.github.io",
    "http://localhost:5173",
    "http://localhost:3000",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:3000",
]


def _is_allowed_origin(origin: str) -> bool:
    """Проверка origin для CORS (в т.ч. варианты с/без слеша и поддомены GitHub Pages)."""
    if not origin or not isinstance(origin, str):
        return False
    origin = origin.rstrip("/")
    if origin in CORS_ORIGINS:
        return True
    if "grakov216500-netizen.github.io" in origin and (origin.startswith("https://") or origin.startswith("http://")):
        return True
    return False


class ForceCORSHeadersMiddleware(BaseHTTPMiddleware):
    """Добавляет CORS-заголовки ко ВСЕМ ответам (в т.ч. 4xx/5xx и OPTIONS)."""
    async def dispatch(self, request: Request, call_next):
        origin = (request.headers.get("origin") or "").strip().rstrip("/") or "https://grakov216500-netizen.github.io"
        if not _is_allowed_origin(origin):
            origin = "https://grakov216500-netizen.github.io"

        def cors_headers():
            return {
                "Access-Control-Allow-Origin": origin,
                "Access-Control-Allow-Credentials": "true",
                "Access-Control-Allow-Methods": "GET, POST, PATCH, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type, Authorization, Accept",
                "Access-Control-Max-Age": "86400",
            }

        # Preflight OPTIONS — сразу 200 с CORS, не передаём в приложение
        if request.method == "OPTIONS":
            from starlette.responses import Response
            return Response(status_code=200, headers=cors_headers())

        try:
            response = await call_next(request)
        except Exception as e:
            from starlette.responses import JSONResponse
            return JSONResponse(
                status_code=500,
                content={"detail": "Internal server error", "error": str(e)},
                headers=cors_headers(),
            )

        for k, v in cors_headers().items():
            response.headers[k] = v
        return response


# Сначала добавляем наш middleware (он выполнится последним при отправке ответа и допишет CORS)
app.add_middleware(ForceCORSHeadersMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# === Путь к БД ===
DB_PATH = os.path.join(os.path.dirname(__file__), "bot.db")

# Токен бота для отправки напоминаний из серверного планировщика (тот же BOT_TOKEN, что и у бота)
BOT_TOKEN = os.getenv("BOT_TOKEN")

# Параметры доступа к Апекс-ВУЗ (используются в apex_parser.py)
APEX_USER = os.getenv("APEX_USER")
APEX_PASS = os.getenv("APEX_PASS")

# Глобальный парсер расписания (ленивая инициализация)
_apex_parser = None


@app.get("/api/health")
async def api_health():
    """Проверка доступности API и CORS (в ответе всегда есть CORS-заголовки)."""
    return {"ok": True, "service": "vitechbot-api"}


@app.on_event("startup")
async def startup_init_db():
    """При старте сервера создаём таблицы и объекты для опроса (только для SQLite)."""
    try:
        import db as db_module
        if not getattr(db_module, "USE_POSTGRES", False):
            import database
            database.DB_NAME = DB_PATH
            database.init_db()
            database.init_survey_objects()
            database.ensure_female_survey_objects()
    except Exception as e:
        print(f"[WARN] Инициализация БД при старте: {e}")

    # Запуск фонового планировщика напоминаний о задачах (не зависит от процесса бота)
    if BOT_TOKEN:
        thread = threading.Thread(target=_task_reminders_loop, daemon=True)
        thread.start()
        print("[OK] Планировщик напоминаний о задачах запущен (каждые 30 сек)")
    else:
        print("[WARN] BOT_TOKEN не задан — напоминания о задачах отправляет только бот")

# === Словарь ролей ===
ROLE_NAMES = {
    'к': 'Курс',
    'дк': 'Дежурный по курсу',
    'с': 'Столовая',
    'дс': 'Дежурный по столовой',
    'ад': 'ГБР',
    'п': 'Патруль',
    'ж': 'Железо',
    'т': 'Тарелки',
    'кпп': 'КПП',
    'гбр': 'ГБР (Группа быстрого реагирования)',
    'зуб': 'ЗУБ',
    'ото': 'ОТО',
    'м': 'Медчасть',
    'путсо': 'ПУТСО',
}

def get_full_role(role_code: str) -> str:
    return ROLE_NAMES.get(role_code.lower(), role_code.upper())


def _fio_match_variants(full_fio: str) -> list:
    """Строит варианты ФИО для сопоставления: полное и в виде инициалов (Граков В.А.)."""
    if not full_fio or not full_fio.strip():
        return []
    parts = [p.strip() for p in full_fio.strip().split() if p.strip()]
    if not parts:
        return []
    variants = [full_fio.strip()]
    if len(parts) >= 3:
        surname, name, patronymic = parts[0], parts[1], parts[2]
        variants.append(f"{surname} {name[0]}.{patronymic[0]}.")
        variants.append(f"{surname} {name[0]}.{patronymic[0]}")
        variants.append(f"{surname} {name[0]} {patronymic[0]}")
        variants.append(f"{surname} {name[0]}. {patronymic[0]}.")
    elif len(parts) == 2:
        variants.append(f"{parts[0]} {parts[1][0]}.")
    return list(dict.fromkeys(variants))


def _get_duty_points_map(conn) -> dict:
    """Роль (код) -> очки по опросу. Основные наряды: Курс, ГБР, Столовая, ЗУБ."""
    role_to_name = {'к': 'Курс', 'гбр': 'ГБР', 'с': 'Столовая', 'зуб': 'ЗУБ'}
    try:
        rows = execute(conn, """
            SELECT o.name, COALESCE(w.weight, 10) as w
            FROM duty_objects o
            LEFT JOIN object_weights w ON o.id = w.object_id
            WHERE o.parent_id IS NULL AND o.name IN ('Курс', 'ГБР', 'Столовая', 'ЗУБ')
        """).fetchall()
        name_to_weight = {r['name']: max(7, min(20, float(r['w'] or 10))) for r in rows}
        return {code: round(name_to_weight.get(name, 10)) for code, name in role_to_name.items()}
    except Exception:
        return {code: 10 for code in role_to_name}


def _create_schedule_notification(conn, group_name: str, enrollment_year: int, title: str, body: str, created_by_telegram_id: int):
    """Создать уведомление об изменении графика для группы (видят все курсанты этой группы)."""
    try:
        execute(conn, """
            INSERT INTO notifications (telegram_id, scope, scope_value, title, body, type, created_by_telegram_id)
            VALUES (NULL, 'group', ?, ?, ?, 'schedule_change', ?)
        """, (group_name or "", title, body or "", created_by_telegram_id))
        conn.commit()
    except Exception as e:
        print(f"[WARN] _create_schedule_notification: {e}")


def _send_telegram_message(chat_id: int, text: str) -> bool:
    """Отправляет сообщение в Telegram через Bot API. Возвращает True при успехе."""
    if not BOT_TOKEN:
        return False
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    body = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "HTML"}).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST", headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        if "bot was blocked" in str(e).lower() or "blocked" in str(e).lower():
            print(f"[REMINDER] Пользователь {chat_id} заблокировал бота")
        else:
            print(f"[REMINDER] Ошибка отправки в {chat_id}: {e}")
        return False


def _run_task_reminders_once():
    """Один проход: найти задачи с дедлайном в окне ±90 сек, отправить напоминания, отметить reminded=1."""
    conn = get_db()
    if not conn:
        return
    try:
        now = datetime.now()
        time_lower = (now - timedelta(seconds=90)).strftime("%Y-%m-%d %H:%M:%S")
        time_upper = (now + timedelta(seconds=90)).strftime("%Y-%m-%d %H:%M:%S")
        rows = execute(conn, """
            SELECT id, text, deadline, user_id FROM tasks
            WHERE done = 0 AND reminded = 0 AND deadline IS NOT NULL
              AND datetime(deadline) >= datetime(?) AND datetime(deadline) <= datetime(?)
        """, (time_lower, time_upper)).fetchall()
        for row in rows:
            task_id = row["id"]
            user_id = row["user_id"]
            text = (row["text"] or "").strip()
            msg = f"⏰ <b>Время выполнить задачу!</b>\n\n{text}"
            if _send_telegram_message(user_id, msg):
                execute(conn,"UPDATE tasks SET reminded = 1 WHERE id = ?", (task_id,))
                conn.commit()
                print(f"[REMINDER] Задача {task_id} → {user_id}")
    except Exception as e:
        print(f"[REMINDER] Ошибка: {e}")
    finally:
        conn.close()


def _task_reminders_loop():
    """Фоновый цикл: каждые 30 сек проверяет дедлайны задач и отправляет напоминания."""
    while True:
        try:
            _run_task_reminders_once()
        except Exception as e:
            print(f"[REMINDER] Цикл: {e}")
        time.sleep(30)

# ============================================
# 1. ПРОФИЛЬ ПОЛЬЗОВАТЕЛЯ (ИСПРАВЛЕНО)
# ============================================
@app.get("/api/user")
async def get_user(telegram_id: int):
    conn = get_db()
    if not conn:
        return {"error": "База данных не найдена"}

    try:
        # --- Узнаём, какие колонки есть в таблице users ---
        cursor = execute(conn,"PRAGMA table_info(users)")
        columns = [row['name'] for row in cursor.fetchall()]
        print(f"[INFO] Колонки в users: {columns}")

        # --- Определяем имя поля с ФИО ---
        if 'full_name' in columns:
            name_col = 'full_name'
        elif 'fio' in columns:
            name_col = 'fio'
        else:
            return {"error": "В таблице users нет поля для ФИО"}

        # --- Определяем имя поля с группой ---
        if 'group_name' in columns:
            group_col = 'group_name'
        elif 'group' in columns:
            group_col = 'group'
        else:
            group_col = None   # необязательное поле

        select_parts = [f"{name_col} as full_name", "enrollment_year"]
        if group_col:
            select_parts.append(f"{group_col} as group_name")
        else:
            select_parts.append("'' as group_name")
        try:
            cursor = execute(conn,"PRAGMA table_info(users)")
            cols = [r['name'] for r in cursor.fetchall()]
            if 'role' in cols:
                select_parts.append("role")
        except Exception:
            pass

        query = f"SELECT {', '.join(select_parts)} FROM users WHERE telegram_id = ?"
        row = execute(conn,query, (telegram_id,)).fetchone()

        if not row:
            return {"error": "Пользователь не найден"}

        user_data = dict(row)
        print(f"[DEBUG] Данные из БД для {telegram_id}: {user_data}")

    except Exception as e:
        print(f"[ERROR] Ошибка запроса пользователя: {e}")
        return {"error": f"Ошибка БД: {str(e)}"}
    finally:
        conn.close()

    try:
        enrollment = int(user_data['enrollment_year'])
        course = get_current_course(enrollment)
    except Exception as e:
        print(f"[ERROR] Ошибка расчёта курса: {e}")
        course = 1

    course_label = "Выпускник" if course >= 5 else str(course)
    out = {
        "full_name": user_data['full_name'],
        "course": str(course),
        "course_label": course_label,
        "enrollment_year": int(user_data.get('enrollment_year', 0)) if user_data.get('enrollment_year') else None,
        "group": user_data.get('group_name', ''),
        "role": user_data.get('role', 'user')
    }
    return out


@app.patch("/api/user")
async def update_user(data: dict):
    """Обновление своего профиля: ФИО, группа, год набора (курс). telegram_id — кто редактирует."""
    telegram_id = data.get("telegram_id")
    fio = data.get("fio")
    group_name = data.get("group_name")
    enrollment_year = data.get("enrollment_year")
    if not telegram_id:
        raise HTTPException(status_code=400, detail="telegram_id обязателен")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        cursor = execute(conn,"PRAGMA table_info(users)")
        cols = [r["name"] for r in cursor.fetchall()]
        updates = []
        params = []
        if fio is not None and str(fio).strip():
            name_col = "fio" if "fio" in cols else "full_name"
            updates.append(f"{name_col} = ?")
            params.append(str(fio).strip())
        if group_name is not None:
            updates.append("group_name = ?")
            params.append(str(group_name).strip() if group_name else "")
        if enrollment_year is not None:
            try:
                y = int(enrollment_year)
                if 2020 <= y <= 2030:
                    updates.append("enrollment_year = ?")
                    params.append(y)
            except (TypeError, ValueError):
                pass
        if not updates:
            conn.close()
            return {"status": "ok"}
        params.append(telegram_id)
        execute(conn,
            f"UPDATE users SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP WHERE telegram_id = ?",
            params
        )
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] update_user: {e}")
        raise HTTPException(status_code=500, detail="Ошибка обновления профиля")
    finally:
        conn.close()


@app.get("/api/profile/duty-stats")
async def get_profile_duty_stats(telegram_id: int):
    """Статистика курсанта: сколько раз болел (замены по болезни + самоотчёты), сколько раз заменял других."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,"SELECT fio FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            conn.close()
            return {"times_sick": 0, "times_replaced": 0}
        fio = (user["fio"] or "").strip()
        times_sick_replace = execute(conn,
            "SELECT COUNT(*) FROM duty_replacements WHERE fio_removed = ? AND reason = 'заболел'",
            (fio,)
        ).fetchone()[0]
        times_sick_self = execute(conn,
            "SELECT COUNT(*) FROM sick_leave_reports WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()[0]
        times_replaced = execute(conn,
            "SELECT COUNT(*) FROM duty_replacements WHERE fio_replacement = ?",
            (fio,)
        ).fetchone()[0]
        conn.close()
        return {"times_sick": times_sick_replace + times_sick_self, "times_replaced": times_replaced}
    except Exception as e:
        print(f"[ERROR] duty-stats: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки статистики")


@app.post("/api/sick-leave/report")
async def report_sick_leave(data: dict):
    """Курсант указывает свой больничный (дата)."""
    telegram_id = data.get("telegram_id")
    report_date = data.get("report_date")
    if not telegram_id or not report_date:
        raise HTTPException(status_code=400, detail="Нужны telegram_id и report_date (YYYY-MM-DD)")
    if len(report_date) != 10 or report_date[4] != "-" or report_date[7] != "-":
        raise HTTPException(status_code=400, detail="Формат даты: YYYY-MM-DD")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        execute(conn,
            "INSERT INTO sick_leave_reports (telegram_id, report_date) VALUES (?, ?)",
            (telegram_id, report_date)
        )
        conn.commit()
        conn.close()
        return {"status": "ok", "message": "Больничный учтён"}
    except Exception as e:
        print(f"[ERROR] sick-leave report: {e}")
        raise HTTPException(status_code=500, detail="Ошибка сохранения")


def _user_role_from_db(telegram_id: int):
    """Роль из БД (admin, assistant, sergeant, user)."""
    conn = get_db()
    if not conn:
        return None
    row = execute(conn,"SELECT role FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    return row["role"] if row else None


@app.get("/api/users")
async def list_users(
    actor_telegram_id: int,
    enrollment_year: int = None,
    group_name: str = None,
    search: str = None
):
    """Список пользователей для админа (все) или помощника (свой курс). Сортировка: курс (год набора), группа, ФИО."""
    role = _user_role_from_db(actor_telegram_id)
    if role not in ("admin", "assistant"):
        raise HTTPException(status_code=403, detail="Доступ только для админа или помощника")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        name_col = "fio" if "fio" in [r["name"] for r in execute(conn,"PRAGMA table_info(users)").fetchall()] else "full_name"
        if role == "assistant":
            row = execute(conn,
                "SELECT enrollment_year, group_name FROM users WHERE telegram_id = ?",
                (actor_telegram_id,)
            ).fetchone()
            if not row:
                conn.close()
                return {"users": []}
            ayear, agroup = row["enrollment_year"], row["group_name"]
            query = f"""
                SELECT telegram_id, {name_col} as fio, group_name, enrollment_year, role
                FROM users WHERE enrollment_year = ? AND status = 'активен'
            """
            params = [ayear]
            if search and search.strip():
                query += f" AND ({name_col} LIKE ?)"
                params.append(f"%{search.strip()}%")
            query += " ORDER BY group_name, fio"
            rows = execute(conn,query, params).fetchall()
        else:
            query = f"""
                SELECT telegram_id, {name_col} as fio, group_name, enrollment_year, role
                FROM users WHERE status = 'активен'
            """
            params = []
            if enrollment_year is not None:
                query += " AND enrollment_year = ?"
                params.append(enrollment_year)
            if group_name and group_name.strip():
                query += " AND group_name = ?"
                params.append(group_name.strip())
            if search and search.strip():
                query += f" AND ({name_col} LIKE ?)"
                params.append(f"%{search.strip()}%")
            query += " ORDER BY enrollment_year DESC, group_name, fio"
            rows = execute(conn,query, params).fetchall()
        users = [
            {
                "telegram_id": r["telegram_id"],
                "fio": r["fio"],
                "group_name": r["group_name"],
                "enrollment_year": r["enrollment_year"],
                "role": r["role"] or "user"
            }
            for r in rows
        ]
        conn.close()
        return {"users": users}
    except Exception as e:
        print(f"[ERROR] list_users: {e}")
        raise HTTPException(status_code=500, detail="Ошибка списка пользователей")


@app.post("/api/users/set-role")
async def set_user_role(data: dict):
    """Назначить/снять роль: admin — помощник или сержант; assistant — только сержант."""
    actor_id = data.get("actor_telegram_id")
    target_id = data.get("target_telegram_id")
    new_role = data.get("role")  # assistant | sergeant | user
    if not actor_id or not target_id or new_role not in ("assistant", "sergeant", "user"):
        raise HTTPException(status_code=400, detail="Нужны actor_telegram_id, target_telegram_id и role (assistant|sergeant|user)")
    actor_role = _user_role_from_db(actor_id)
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    if actor_role == "admin":
        pass  # может назначать без ограничений
    elif actor_role == "assistant":
        a_row = execute(conn,"SELECT enrollment_year FROM users WHERE telegram_id = ?", (actor_id,)).fetchone()
        t_row = execute(conn,"SELECT enrollment_year, group_name FROM users WHERE telegram_id = ?", (target_id,)).fetchone()
        if not a_row or not t_row or a_row["enrollment_year"] != t_row["enrollment_year"]:
            conn.close()
            raise HTTPException(status_code=403, detail="Можно менять только пользователей своего курса")
        if new_role == "assistant":
            cnt = execute(conn,
                "SELECT COUNT(*) FROM users WHERE enrollment_year = ? AND role = 'assistant'",
                (t_row["enrollment_year"],)
            ).fetchone()[0]
            if cnt >= 6:
                conn.close()
                raise HTTPException(status_code=403, detail="На курсе уже 6 помощников (лимит)")
        elif new_role == "sergeant":
            grp = t_row["group_name"] or ""
            cnt = execute(conn,
                "SELECT COUNT(*) FROM users WHERE group_name = ? AND role = 'sergeant'",
                (grp,)
            ).fetchone()[0]
            if cnt >= 4:
                conn.close()
                raise HTTPException(status_code=403, detail="В группе уже 4 сержанта (лимит)")
    else:
        conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    try:
        execute(conn,"UPDATE users SET role = ?, updated_at = CURRENT_TIMESTAMP WHERE telegram_id = ?", (new_role, target_id))
        conn.commit()
        conn.close()
        return {"status": "ok", "role": new_role}
    except Exception as e:
        print(f"[ERROR] set_user_role: {e}")
        raise HTTPException(status_code=500, detail="Ошибка назначения роли")


# ============================================
# 2. НАРЯДЫ ПОЛЬЗОВАТЕЛЯ
# ============================================
@app.get("/api/duties")
async def get_duties(telegram_id: int, month: str = None, year: int = None):
    """
    Получает наряды пользователя.
    Если указаны month и year - возвращает наряды за конкретный месяц.
    Иначе возвращает все наряды и ближайший.
    """
    conn = get_db()
    if not conn:
        return {"error": "База данных не найдена"}

    try:
        # --- Получаем user_id и ФИО ---
        user = execute(conn,
            "SELECT id, fio FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
    except Exception as e:
        print(f"[ERROR] Ошибка при поиске user_id: {e}")
        return {"error": f"Ошибка БД: {str(e)}"}

    if not user:
        conn.close()
        return {"error": "Пользователь не найден"}

    user_id = user['id']
    user_fio = user['fio']  # ФИО из users — в Excel (duty_schedule) должно быть то же написание

    try:
        # Проверяем, какая таблица используется
        # Сначала пробуем duty_schedule (новая структура)
        try:
            cursor = execute(conn,"PRAGMA table_info(duty_schedule)")
            schedule_columns = [row['name'] for row in cursor.fetchall()]
            if schedule_columns:
                # Используем duty_schedule
                if month is not None and year is not None:
                    # Фильтр по месяцу (month/year могут прийти строками из query)
                    try:
                        m = int(month)
                        y = int(year)
                    except (TypeError, ValueError):
                        m, y = None, None
                    if m is not None and y is not None and 1 <= m <= 12:
                        month_start = f"{y}-{m:02d}-01"
                        if m == 12:
                            month_end = f"{y + 1}-01-01"
                        else:
                            month_end = f"{y}-{m + 1:02d}-01"
                    else:
                        month_start = month_end = None
                    if month_start and month_end:
                        fio_variants = _fio_match_variants(user_fio) or [user_fio or ""]
                        placeholders = ",".join(["?"] * len(fio_variants))
                        query = f"""
                            SELECT date, role, group_name, enrollment_year
                            FROM duty_schedule
                            WHERE fio IN ({placeholders}) AND date >= ? AND date < ?
                            ORDER BY date
                        """
                        rows = execute(conn,query, (*fio_variants, month_start, month_end)).fetchall()
                    else:
                        fio_variants = _fio_match_variants(user_fio) or [user_fio or ""]
                        placeholders = ",".join(["?"] * len(fio_variants))
                        rows = execute(conn,
                            f"SELECT date, role, group_name, enrollment_year FROM duty_schedule WHERE fio IN ({placeholders}) ORDER BY date",
                            tuple(fio_variants),
                        ).fetchall()
                else:
                    fio_variants = _fio_match_variants(user_fio) or [user_fio or ""]
                    placeholders = ",".join(["?"] * len(fio_variants))
                    query = f"""
                        SELECT date, role, group_name, enrollment_year
                        FROM duty_schedule
                        WHERE fio IN ({placeholders})
                        ORDER BY date
                    """
                    rows = execute(conn,query, tuple(fio_variants)).fetchall()
                
                duties_list = []
                points_map = _get_duty_points_map(conn)
                for row in rows:
                    # Получаем участников наряда на эту дату
                    partners_query = """
                        SELECT fio, group_name
                        FROM duty_schedule
                        WHERE date = ? AND role = ? AND enrollment_year = ?
                        ORDER BY group_name, fio
                    """
                    partners = execute(conn,partners_query, (row['date'], row['role'], row['enrollment_year'])).fetchall()
                    role_lower = (row['role'] or '').strip().lower()
                    points = points_map.get(role_lower, 10)
                    duties_list.append({
                        "date": row['date'],
                        "role": row['role'],
                        "role_full": get_full_role(row['role']),
                        "group": row['group_name'],
                        "partners": [{"fio": p['fio'], "group": p['group_name']} for p in partners],
                        "points": points,
                    })
                
                conn.close()
                
                # Ближайший наряд (если не указан месяц)
                if not month:
                    today = datetime.now().strftime("%Y-%m-%d")
                    upcoming = [d for d in duties_list if d['date'] >= today]
                    next_duty = upcoming[0] if upcoming else None
                else:
                    next_duty = None
                
                return {
                    "duties": duties_list,
                    "next_duty": next_duty,
                    "total": len(duties_list)
                }
        except Exception:
            pass

        # Если duty_schedule не работает, пробуем duties (старая структура)
        try:
            cursor = execute(conn,"PRAGMA table_info(duties)")
            columns = [row['name'] for row in cursor.fetchall()]
        except Exception:
            conn.close()
            return {
                "duties": [],
                "next_duty": None,
                "total": 0,
                "error": "График нарядов ещё не загружен. Когда данные появятся, здесь отобразятся ваши наряды."
            }

        if columns and 'object_type' in columns:
            query = """
                SELECT date, role, object_type
                FROM duties
                WHERE user_id = ?
                ORDER BY date
            """
            rows = execute(conn,query, (user_id,)).fetchall()
            duties_list = [
                {
                    "date": row['date'],
                    "role": row['role'],
                    "role_full": get_full_role(row['role']),
                    "object": row['object_type'] or "—",
                    "partners": []  # Старая структура не поддерживает участников
                }
                for row in rows
            ]
        else:
            query = """
                SELECT date, role
                FROM duties
                WHERE user_id = ?
                ORDER BY date
            """
            rows = execute(conn,query, (user_id,)).fetchall()
            duties_list = [
                {
                    "date": row['date'],
                    "role": row['role'],
                    "role_full": get_full_role(row['role']),
                    "object": "—",
                    "partners": []
                }
                for row in rows
            ]

    except Exception as e:
        print(f"[ERROR] Ошибка при запросе нарядов: {e}")
        err_msg = str(e).lower()
        if "no such table" in err_msg or "duties" in err_msg:
            conn.close()
            return {
                "duties": [],
                "next_duty": None,
                "total": 0,
                "error": "График нарядов ещё не загружен. Когда данные появятся, здесь отобразятся ваши наряды."
            }
        return {"error": f"Ошибка БД: {str(e)}"}
    finally:
        conn.close()

    # --- Ближайший наряд ---
    today = datetime.now().strftime("%Y-%m-%d")
    upcoming = [d for d in duties_list if d['date'] >= today]
    next_duty = upcoming[0] if upcoming else None

    return {
        "duties": duties_list,
        "next_duty": next_duty,
        "total": len(duties_list)
    }


@app.get("/api/duties/by-date")
async def get_duties_by_date(date: str, telegram_id: int = 0):
    """
    Возвращает всех участников наряда на конкретную дату.
    Если telegram_id передан — фильтрует по курсу (enrollment_year) пользователя.
    """
    conn = get_db()
    if not conn:
        return {"error": "База данных не найдена"}
    
    try:
        ey = None
        if telegram_id:
            user = execute(conn,"SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
            if user:
                ey = user["enrollment_year"]

        if ey:
            query = """
                SELECT fio, role, group_name, enrollment_year, gender
                FROM duty_schedule
                WHERE date = ? AND enrollment_year = ?
                ORDER BY role, group_name, fio
            """
            rows = execute(conn,query, (date, ey)).fetchall()
        else:
            query = """
                SELECT fio, role, group_name, enrollment_year, gender
                FROM duty_schedule
                WHERE date = ?
                ORDER BY role, group_name, fio
            """
            rows = execute(conn,query, (date,)).fetchall()
        
        by_role = {}
        for row in rows:
            role = row['role']
            if role not in by_role:
                by_role[role] = []
            by_role[role].append({
                "fio": row['fio'],
                "group": row['group_name'],
                "course": row['enrollment_year'],
                "gender": row['gender']
            })
        
        return {
            "date": date,
            "by_role": by_role,
            "total": len(rows)
        }
    except Exception as e:
        print(f"[ERROR] Ошибка при запросе нарядов по дате: {e}")
        return {"error": f"Ошибка БД: {str(e)}"}
    finally:
        if conn:
            conn.close()

@app.get("/api/duties/available-months")
async def get_available_months(telegram_id: int):
    """Возвращает список месяцев (YYYY-MM), для которых есть загруженные графики в рамках курса пользователя."""
    conn = get_db()
    if not conn:
        return {"months": []}
    try:
        user = execute(conn,
            "SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user:
            return {"months": []}
        ey = user["enrollment_year"]
        rows = execute(conn, """
            SELECT DISTINCT substr(date, 1, 7) as ym
            FROM duty_schedule
            WHERE enrollment_year = ?
            ORDER BY ym
        """, (ey,)).fetchall()
        return {"months": [r["ym"] for r in rows]}
    except Exception as e:
        print(f"[ERROR] available-months: {e}")
        return {"months": []}
    finally:
        conn.close()


@app.get("/api/duties/day-detail")
async def get_duty_day_detail(date: str, role: str, telegram_id: int):
    """Подробная информация о конкретном наряде (роль) на конкретную дату: все участники того же курса."""
    conn = get_db()
    if not conn:
        return {"error": "БД не найдена"}
    try:
        user = execute(conn,
            "SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user:
            return {"error": "Пользователь не найден"}
        ey = user["enrollment_year"]
        rows = execute(conn, """
            SELECT ds.fio, ds.group_name, ds.gender, u.telegram_id
            FROM duty_schedule ds
            LEFT JOIN users u ON u.fio = ds.fio AND u.status = 'активен'
            WHERE ds.date = ? AND ds.role = ? AND ds.enrollment_year = ?
            ORDER BY ds.group_name, ds.fio
        """, (date, role, ey)).fetchall()
        # Подставляем telegram_id по инициалам, если точное ФИО не совпало
        fio_to_telegram = {}
        for r in execute(conn,
            "SELECT fio, telegram_id FROM users WHERE enrollment_year = ? AND status = 'активен'", (ey,)
        ).fetchall():
            for v in _fio_match_variants(r["fio"]) or [r["fio"]]:
                fio_to_telegram[v] = r["telegram_id"]
        participants = []
        for r in rows:
            tid = r["telegram_id"] if r["telegram_id"] is not None else fio_to_telegram.get(r["fio"])
            participants.append({
                "fio": r["fio"], "group": r["group_name"], "gender": r["gender"], "telegram_id": tid
            })
        
        shift_data = []
        canteen_data = []

        # Автораспределение за 3 часа до наряда: если назначений ещё нет.
        try:
            duty_start = datetime.strptime(date, "%Y-%m-%d").replace(hour=18, minute=30, second=0, microsecond=0)
            auto_time = duty_start - timedelta(hours=3)  # 15:30 того же дня
            now = datetime.now()
        except Exception:
            duty_start = None
            auto_time = None
            now = datetime.now()

        if role in ("к", "гбр"):
            try:
                s_rows = execute(conn, """
                    SELECT fio, shift FROM duty_shift_assignments
                    WHERE date = ? AND role = ? AND enrollment_year = ?
                    ORDER BY shift, fio
                """, (date, role, ey)).fetchall()
                if not s_rows and auto_time and now >= auto_time:
                    # Автоматически распределяем смены, если ещё не распределены
                    try:
                        distribute_shifts_for_date(date, role, ey, conn)
                    except Exception:
                        pass
                    s_rows = execute(conn, """
                        SELECT fio, shift FROM duty_shift_assignments
                        WHERE date = ? AND role = ? AND enrollment_year = ?
                        ORDER BY shift, fio
                    """, (date, role, ey)).fetchall()
                shift_data = [{"fio": r["fio"], "shift": r["shift"]} for r in s_rows]
            except Exception:
                pass
        elif role == "с":
            try:
                c_rows = execute(conn, """
                    SELECT fio, object_name FROM duty_canteen_assignments
                    WHERE date = ? AND enrollment_year = ?
                    ORDER BY object_name, fio
                """, (date, ey)).fetchall()
                if not c_rows and auto_time and now >= auto_time:
                    try:
                        distribute_canteen_for_date(date, ey, conn)
                    except Exception:
                        pass
                    c_rows = execute(conn, """
                        SELECT fio, object_name FROM duty_canteen_assignments
                        WHERE date = ? AND enrollment_year = ?
                        ORDER BY object_name, fio
                    """, (date, ey)).fetchall()
                canteen_data = [{"fio": r["fio"], "object": r["object_name"]} for r in c_rows]
            except Exception:
                pass
        
        return {
            "date": date,
            "role": role,
            "role_full": get_full_role(role),
            "participants": participants,
            "count": len(participants),
            "shifts": shift_data,
            "canteen": canteen_data,
        }
    except Exception as e:
        print(f"[ERROR] day-detail: {e}")
        return {"error": str(e)}
    finally:
        conn.close()


# ============================================
# 2.5. РАСПРЕДЕЛЕНИЕ ПО СМЕНАМ И ОБЪЕКТАМ
# ============================================

CANTEEN_OBJECTS = ["ГЦ", "овощи", "тарелки", "железо", "стаканы", "лента"]

def distribute_shifts_for_date(date_str: str, role: str, ey: int, conn):
    """Распределяет людей по сменам для Курс/ГБР. Возвращает список назначений."""
    rows = execute(conn, """
        SELECT fio, group_name FROM duty_schedule
        WHERE date = ? AND role = ? AND enrollment_year = ?
        ORDER BY fio
    """, (date_str, role, ey)).fetchall()
    people = [r["fio"] for r in rows]
    if not people:
        return []
    
    random.shuffle(people)
    assignments = []
    
    if role == "к":
        # Фиксированный дежурный по курсу (shift=0) + до 3 дневальных (shift=1..3)
        if len(people) == 1:
            assignments.append({"fio": people[0], "shift": 0})
        else:
            # Первый — дежурный по курсу
            assignments.append({"fio": people[0], "shift": 0})
            day_count = min(3, len(people) - 1)
            for i in range(day_count):
                assignments.append({"fio": people[1 + i], "shift": i + 1})
            # Остальные, если есть, остаются без смены (shift=0)
            for fio in people[1 + day_count:]:
                assignments.append({"fio": fio, "shift": 0})
    elif role == "гбр":
        for i, fio in enumerate(people):
            shift = (i // 2) + 1
            assignments.append({"fio": fio, "shift": shift})
    else:
        for i, fio in enumerate(people):
            shift = (i % 3) + 1
            assignments.append({"fio": fio, "shift": shift})
    
    execute(conn,"DELETE FROM duty_shift_assignments WHERE date = ? AND role = ? AND enrollment_year = ?",
                 (date_str, role, ey))
    for a in assignments:
        execute(conn, """
            INSERT OR REPLACE INTO duty_shift_assignments (date, role, fio, shift, enrollment_year)
            VALUES (?, ?, ?, ?, ?)
        """, (date_str, role, a["fio"], a["shift"], ey))
        execute(conn, """
            INSERT INTO duty_assignment_history (fio, date, role, shift, enrollment_year)
            VALUES (?, ?, ?, ?, ?)
        """, (a["fio"], date_str, role, a["shift"], ey))
    conn.commit()
    return assignments


def distribute_canteen_for_date(date_str: str, ey: int, conn):
    """Распределяет людей по объектам столовой с учётом рейтинга и истории."""
    rows = execute(conn, """
        SELECT fio, group_name FROM duty_schedule
        WHERE date = ? AND role = 'с' AND enrollment_year = ?
        ORDER BY fio
    """, (date_str, ey)).fetchall()
    people = [r["fio"] for r in rows]
    if not people:
        return []

    scores = {}
    for fio in people:
        user_row = execute(conn,"SELECT global_score FROM users WHERE fio = ? AND enrollment_year = ?",
                                (fio, ey)).fetchone()
        gs = user_row["global_score"] if user_row and user_row["global_score"] else 0
        
        hist = execute(conn, """
            SELECT sub_object FROM duty_assignment_history
            WHERE fio = ? AND role = 'с' AND enrollment_year = ?
            ORDER BY date DESC LIMIT 5
        """, (fio, ey)).fetchall()
        history = [h["sub_object"] for h in hist if h["sub_object"]]
        
        streak_penalty = 0
        if len(history) >= 2:
            weights_map = {}
            try:
                w_rows = execute(conn, """
                    SELECT do.name, ow.weight FROM duty_objects do
                    JOIN object_weights ow ON do.id = ow.object_id
                """).fetchall()
                weights_map = {r["name"]: r["weight"] for r in w_rows}
            except Exception:
                pass
            
            heavy = [o for o in CANTEEN_OBJECTS if weights_map.get(o, 10) >= 12]
            if history[0] in heavy and len(history) > 1 and history[1] in heavy:
                streak_penalty = 5
        
        scores[fio] = 0.5 * gs + streak_penalty
    
    sorted_people = sorted(people, key=lambda f: scores.get(f, 0))
    
    weights_map = {}
    try:
        w_rows = execute(conn, """
            SELECT do.name, ow.weight FROM duty_objects do
            JOIN object_weights ow ON do.id = ow.object_id
        """).fetchall()
        weights_map = {r["name"]: r["weight"] for r in w_rows}
    except Exception:
        pass
    
    objects_sorted = sorted(CANTEEN_OBJECTS, key=lambda o: weights_map.get(o, 10), reverse=True)
    
    assignments = []
    obj_idx = 0
    for fio in sorted_people:
        obj = objects_sorted[obj_idx % len(objects_sorted)]
        obj_idx += 1
        assignments.append({"fio": fio, "object": obj})
    
    execute(conn,"DELETE FROM duty_canteen_assignments WHERE date = ? AND enrollment_year = ?",
                 (date_str, ey))
    for a in assignments:
        execute(conn, """
            INSERT OR REPLACE INTO duty_canteen_assignments (date, fio, object_name, enrollment_year)
            VALUES (?, ?, ?, ?)
        """, (date_str, a["fio"], a["object"], ey))
        execute(conn, """
            INSERT INTO duty_assignment_history (fio, date, role, sub_object, enrollment_year)
            VALUES (?, ?, 'с', ?, ?)
        """, (a["fio"], date_str, a["object"], ey))
    conn.commit()
    return assignments


@app.post("/api/duties/distribute")
async def distribute_duty(date: str = Form(...), role: str = Form(...), telegram_id: int = Form(...)):
    """Ручной запуск распределения по сменам/объектам (для сержантов/админов)."""
    conn = get_db()
    if not conn:
        raise HTTPException(500, detail="БД не найдена")
    try:
        user = execute(conn,"SELECT enrollment_year, role as user_role FROM users WHERE telegram_id = ?",
                            (telegram_id,)).fetchone()
        if not user:
            raise HTTPException(404, detail="Пользователь не найден")
        if user["user_role"] not in ("admin", "sergeant", "assistant"):
            raise HTTPException(403, detail="Недостаточно прав")
        ey = user["enrollment_year"]
        
        if role == "с":
            result = distribute_canteen_for_date(date, ey, conn)
            return {"status": "ok", "type": "canteen", "assignments": result, "count": len(result)}
        else:
            result = distribute_shifts_for_date(date, role, ey, conn)
            return {"status": "ok", "type": "shift", "assignments": result, "count": len(result)}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] distribute: {e}")
        raise HTTPException(500, detail=str(e))
    finally:
        conn.close()


@app.get("/api/duties/shifts")
async def get_duty_shifts(date: str, role: str, telegram_id: int):
    """Получить распределение по сменам на дату."""
    conn = get_db()
    if not conn:
        return {"assignments": []}
    try:
        user = execute(conn,"SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            return {"assignments": []}
        ey = user["enrollment_year"]
        rows = execute(conn, """
            SELECT fio, shift FROM duty_shift_assignments
            WHERE date = ? AND role = ? AND enrollment_year = ?
            ORDER BY shift, fio
        """, (date, role, ey)).fetchall()
        return {"date": date, "role": role, "assignments": [{"fio": r["fio"], "shift": r["shift"]} for r in rows]}
    except Exception as e:
        return {"assignments": [], "error": str(e)}
    finally:
        conn.close()


@app.get("/api/duties/canteen-assignments")
async def get_canteen_assignments(date: str, telegram_id: int):
    """Получить распределение по объектам столовой."""
    conn = get_db()
    if not conn:
        return {"assignments": []}
    try:
        user = execute(conn,"SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            return {"assignments": []}
        ey = user["enrollment_year"]
        rows = execute(conn, """
            SELECT fio, object_name FROM duty_canteen_assignments
            WHERE date = ? AND enrollment_year = ?
            ORDER BY object_name, fio
        """, (date, ey)).fetchall()
        return {"date": date, "assignments": [{"fio": r["fio"], "object": r["object_name"]} for r in rows]}
    except Exception as e:
        return {"assignments": [], "error": str(e)}
    finally:
        conn.close()


# ============================================
# 3. ЗАДАЧНИК: API ДЛЯ WEBAPP
# ============================================

@app.get("/api/tasks")
async def get_tasks(user_id: int):
    conn = get_db()
    if not conn:
        return {"error": "База данных не найдена"}

    try:
        # Проверим, есть ли таблица tasks
        cursor = execute(conn,"PRAGMA table_info(tasks)")
        columns = [row['name'] for row in cursor.fetchall()]
        if not columns:
            print("[ERROR] Таблица tasks не найдена")
            return {"error": "Таблица задач не найдена"}

        # Выбираем задачи
        query = """
            SELECT id, text, done, deadline 
            FROM tasks 
            WHERE user_id = ? 
            ORDER BY done, deadline IS NULL, deadline
        """
        rows = execute(conn,query, (user_id,)).fetchall()
        tasks = []
        for row in rows:
            task = dict(row)
            # Форматируем deadline для отображения
            if task['deadline']:
                try:
                    dt = datetime.fromisoformat(task['deadline'])
                    task['formatted_deadline'] = dt.strftime('%d %H:%M')
                except:
                    task['formatted_deadline'] = None
            else:
                task['formatted_deadline'] = None
            tasks.append(task)
        return tasks

    except Exception as e:
        print(f"[ERROR] Ошибка загрузки задач: {e}")
        return {"error": f"Ошибка БД: {str(e)}"}
    finally:
        conn.close()


@app.post("/api/add_task")
async def add_task(data: dict):
    user_id = data.get('user_id')
    text = data.get('text')

    if not user_id or not text:
        raise HTTPException(status_code=400, detail="user_id и text обязательны")

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        execute(conn, """
            INSERT INTO tasks (user_id, text, done, reminded, deadline) 
            VALUES (?, ?, 0, 0, NULL)
        """, (user_id, text.strip()))
        conn.commit()
        print(f"✅ Добавлена задача: '{text}' для user_id={user_id}")
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] Ошибка добавления задачи: {e}")
        raise HTTPException(status_code=500, detail="Не удалось добавить задачу")
    finally:
        conn.close()


@app.post("/api/done_task")
async def done_task(data: dict):
    task_id = data.get('task_id')
    user_id = data.get('user_id')
    done = data.get('done', True)

    if not task_id or not user_id:
        raise HTTPException(status_code=400, detail="task_id и user_id обязательны")

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        execute(conn,"UPDATE tasks SET done = ? WHERE id = ? AND user_id = ?", (int(done), task_id, user_id))
        conn.commit()
        action = "выполнена" if done else "восстановлена"
        print(f"✅ Задача {task_id} отмечена как {action}")
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] Ошибка обновления статуса задачи: {e}")
        raise HTTPException(status_code=500, detail="Не удалось обновить задачу")
    finally:
        conn.close()


@app.post("/api/delete_task")
async def delete_task(data: dict):
    task_id = data.get('task_id')
    user_id = data.get('user_id')

    if not task_id or not user_id:
        raise HTTPException(status_code=400, detail="task_id и user_id обязательны")

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        execute(conn,"DELETE FROM tasks WHERE id = ? AND user_id = ?", (task_id, user_id))
        conn.commit()
        print(f"✅ Задача {task_id} удалена")
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] Ошибка удаления задачи: {e}")
        raise HTTPException(status_code=500, detail="Не удалось удалить задачу")
    finally:
        conn.close()


@app.post("/api/edit_task")
async def edit_task(data: dict):
    task_id = data.get('task_id')
    text = data.get('text')
    user_id = data.get('user_id')

    if not task_id or not text or not user_id:
        raise HTTPException(status_code=400, detail="task_id, text и user_id обязательны")

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        execute(conn,"UPDATE tasks SET text = ? WHERE id = ? AND user_id = ?", (text.strip(), task_id, user_id))
        conn.commit()
        print(f"✅ Задача {task_id} отредактирована: '{text}'")
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] Ошибка редактирования задачи: {e}")
        raise HTTPException(status_code=500, detail="Не удалось редактировать задачу")
    finally:
        conn.close()


@app.post("/api/set_reminder")
async def set_reminder(data: dict):
    task_id = data.get('task_id')
    deadline = data.get('deadline')  # 'YYYY-MM-DD HH:MM:SS'
    user_id = data.get('user_id')

    if not task_id or not deadline or not user_id:
        raise HTTPException(status_code=400, detail="task_id, deadline и user_id обязательны")

    # Валидация формата
    try:
        datetime.fromisoformat(deadline)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD HH:MM:SS")

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        execute(conn, """
            UPDATE tasks 
            SET deadline = ?, reminded = 0 
            WHERE id = ? AND user_id = ?
        """, (deadline, task_id, user_id))
        conn.commit()
        print(f"✅ Напоминание установлено: задача {task_id} → {deadline}")
        return {"status": "ok"}
    except Exception as e:
        print(f"[ERROR] Ошибка установки напоминания: {e}")
        raise HTTPException(status_code=500, detail="Не удалось установить напоминание")
    finally:
        conn.close()


# ============================================
# 4. РАСПИСАНИЕ (интеграция с Апекс)
# ============================================


def _get_apex_parser():
    """Ленивая инициализация парсера расписания Апекс-ВУЗ."""
    global _apex_parser
    if _apex_parser is None:
        if not APEX_USER or not APEX_PASS:
            raise HTTPException(
                status_code=503,
                detail="APEX_USER / APEX_PASS не заданы на сервере"
            )
        try:
            from apex_parser import create_default_parser
            _apex_parser = create_default_parser()
        except Exception as e:
            print(f"[ERROR] Инициализация ApexScheduleParser: {e}")
            raise HTTPException(status_code=500, detail="Ошибка инициализации парсера расписания")
    return _apex_parser


@app.get("/api/schedule/today")
async def get_today_schedule(telegram_id: int, date: str = None):
    """
    Расписание на день для пользователя.
    date: YYYY-MM-DD (опционально). Если не передан — текущая дата.
    Если сегодня сб/вс, фронт может передать date следующего понедельника.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT fio, group_name, enrollment_year FROM users WHERE telegram_id = ?",
            (telegram_id,),
        ).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")

        group_name = (user["group_name"] or "").strip()
        year = user["enrollment_year"]
        if not group_name:
            raise HTTPException(status_code=400, detail="У пользователя не указана группа")

        target_date = datetime.now().date()
        if date:
            try:
                target_date = datetime.strptime(date.strip()[:10], "%Y-%m-%d").date()
            except ValueError:
                pass

        lessons = []
        message = None
        try:
            parser = _get_apex_parser()
            lessons = parser.get_schedule_for_date(group_name, year, target_date)
        except HTTPException as he:
            if he.status_code == 503:
                message = "Сервис расписания временно недоступен (APEX не настроен)"
            else:
                raise
        except ValueError as ve:
            print(f"[WARN] Расписание Апекс (группа не найдена): {ve}")
            message = "Группа не найдена в Апексе"
        except Exception as e:
            print(f"[WARN] Расписание Апекс: {e}")
            message = "Не удалось загрузить расписание (выходной или сайт недоступен)"

        return {
            "date": target_date.strftime("%Y-%m-%d"),
            "group": group_name,
            "year": year,
            "lessons": lessons,
            "message": message,
        }
    finally:
        conn.close()


@app.get("/api/schedule/week")
async def get_week_schedule(telegram_id: int, date: str = None):
    """
    Расписание на учебную неделю (Пн–Пт).
    date: YYYY-MM-DD — любой день недели; по нему определяется понедельник.
    Если не передан — текущая неделя (или следующая, если сейчас сб/вс).
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT fio, group_name, enrollment_year FROM users WHERE telegram_id = ?",
            (telegram_id,),
        ).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        group_name = (user["group_name"] or "").strip()
        year = user["enrollment_year"]
        if not group_name:
            raise HTTPException(status_code=400, detail="У пользователя не указана группа")

        target = datetime.now().date()
        if date:
            try:
                target = datetime.strptime(date.strip()[:10], "%Y-%m-%d").date()
            except ValueError:
                pass
        # Понедельник = target - (weekday-1), где weekday 0=Mon, 6=Sun
        weekday = target.weekday()
        monday = target - timedelta(days=weekday)
        week_start = monday

        week_schedule = {}
        message = None
        try:
            parser = _get_apex_parser()
            week_schedule = parser.get_schedule_for_week(group_name, year, week_start)
        except HTTPException as he:
            if he.status_code == 503:
                message = "Сервис расписания временно недоступен (APEX не настроен)"
            else:
                raise
        except ValueError as ve:
            print(f"[WARN] Расписание Апекс (группа не найдена): {ve}")
            message = "Группа не найдена в Апексе"
        except Exception as e:
            print(f"[WARN] Расписание Апекс (неделя): {e}")
            message = "Не удалось загрузить расписание"

        return {
            "week_start": week_start.strftime("%Y-%m-%d"),
            "group": group_name,
            "year": year,
            "schedule": week_schedule,
            "message": message,
        }
    finally:
        conn.close()


@app.get("/api/schedule/check_month")
async def check_schedule_month(group: str, enrollment_year: int, month: str):
    """Проверка: есть ли уже график за указанный месяц для группы. month = YYYY-MM."""
    if len(month) != 7 or month[4] != "-":
        raise HTTPException(status_code=400, detail="month должен быть в формате YYYY-MM")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        start = month + "-01"
        try:
            dt = datetime.strptime(month, "%Y-%m")
            if dt.month == 12:
                end = f"{dt.year + 1}-01-01"
            else:
                end = f"{dt.year}-{dt.month + 1:02d}-01"
        except Exception:
            end = start
        row = execute(conn, """
            SELECT 1 FROM duty_schedule
            WHERE group_name = ? AND enrollment_year = ?
            AND date >= ? AND date < ?
            LIMIT 1
        """, (group, enrollment_year, start, end)).fetchone()
        return {"has_data": row is not None, "month": month, "group": group}
    finally:
        conn.close()


# Путь к шаблону графика: из env или файл в корне проекта
SCHEDULE_TEMPLATE_PATH = os.environ.get(
    "SCHEDULE_TEMPLATE_PATH",
    os.path.join(os.path.dirname(__file__), "graph_ИО6 — копия.xlsx")
)


def _generate_schedule_template_bytes():
    """
    Возвращает .xlsx шаблон графика нарядов.
    В первую очередь пытаемся отдать пользовательский файл SCHEDULE_TEMPLATE_PATH,
    чтобы курсанты скачивали именно ваш актуальный шаблон.
    Если файла нет или произошла ошибка чтения — генерируем запасной шаблон
    Если файла нет или произошла ошибка чтения — генерируем запасной шаблон
    тем же форматом, который ожидает парсер (группа E1, год AO4, ФИО F6:H55,
    месяц I4, дни I5:AM5, ячейки I6:AM55).
    """
    # 1. Пользовательский шаблон с диска
    try:
        if os.path.exists(SCHEDULE_TEMPLATE_PATH):
            with open(SCHEDULE_TEMPLATE_PATH, "rb") as f:
                return f.read()
    except Exception as e:
        print(f"[WARN] Не удалось прочитать пользовательский шаблон: {e}")

    # 2. Резервный генератор (старый вариант через openpyxl)
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws.title == "Sheet":
        ws.title = "График"
    # Группа — E1
    ws["E1"] = "Группа (напр. ИО61)"
    ws["E2"] = ""
    # Год — AO4 (клише =$AO$4: в этой ячейке год, чтобы графики не терялись)
    from datetime import date as _date
    ws["AO4"] = _date.today().year
    # Заголовки месяц/дни — I4:AM5
    ws.cell(4, 9, "месяц (напр. январь)")
    for c in range(1, 32):
        ws.cell(5, 8 + c, c)
    # ФИО — колонки F,G,H строки 6-21
    ws.cell(6, 6, "Фамилия")
    ws.cell(6, 7, "Имя")
    ws.cell(6, 8, "Отчество")
    for r in range(7, 52):
        for c in range(6, 9):
            ws.cell(r, c, "")
    ws.cell(52, 1, "Роли: к, дк, с, ад, гбр, зуб, столовая, ото, м, путсо и т.д.")

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.get("/api/schedule/template")
async def get_schedule_template(telegram_id: int = None):
    """Скачать шаблон .xlsx. Если передан telegram_id и есть шаблон для группы пользователя — отдаём его."""
    group_template_path = None
    if telegram_id:
        conn = get_db()
        if conn:
            try:
                row = execute(conn, "SELECT group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
                if row:
                    safe_group = (row["group_name"] or "").replace("/", "_").strip() or "group"
                    year = row["enrollment_year"] or ""
                    templates_dir = os.path.join(os.path.dirname(__file__), "group_templates")
                    group_template_path = os.path.join(templates_dir, f"{safe_group}_{year}.xlsx")
                    if not os.path.isfile(group_template_path):
                        group_template_path = None
            finally:
                conn.close()
    if group_template_path:
        try:
            with open(group_template_path, "rb") as f:
                data = f.read()
            return Response(
                content=data,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=schedule_template.xlsx"}
            )
        except Exception as e:
            print(f"[WARN] Шаблон группы не прочитан: {e}")
    data = _generate_schedule_template_bytes()
    if not data:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule_template.xlsx"}
    )


@app.post("/api/schedule/upload-template")
async def upload_group_template(
    file: UploadFile = File(...),
    telegram_id: int = Form(...),
):
    """Сержант/помощник/админ загружает шаблон для своей группы. Подменяет общий шаблон при скачивании для этой группы."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Нужен файл .xlsx")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        row = execute(conn, "SELECT group_name, enrollment_year, role FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not row or row["role"] not in ("sergeant", "assistant", "admin"):
            raise HTTPException(status_code=403, detail="Только сержант своей группы, помощник или админ могут загрузить шаблон для группы")
        safe_group = (row["group_name"] or "").replace("/", "_").strip() or "group"
        year = row["enrollment_year"] or ""
        templates_dir = os.path.join(os.path.dirname(__file__), "group_templates")
        os.makedirs(templates_dir, exist_ok=True)
        path = os.path.join(templates_dir, f"{safe_group}_{year}.xlsx")
        contents = await file.read()
        with open(path, "wb") as f:
            f.write(contents)
        conn.close()
        return {"status": "ok", "message": "Шаблон для группы сохранён. При скачивании шаблона курсанты вашей группы получат этот файл."}
    except HTTPException:
        raise
    finally:
        if conn:
            conn.close()


@app.post("/api/schedule/upload")
async def upload_schedule(
    file: UploadFile = File(...),
    telegram_id: int = Form(...),
    overwrite: int = Form(0),
):
    """Загрузка графика из .xlsx. Доступно сержанту (своя группа), помощнику/админу."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Нужен файл .xlsx")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            raise HTTPException(status_code=403, detail="Нет прав на загрузку графика")
    finally:
        conn.close()

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")
    tmp = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(content)
            tmp = f.name
        from utils.parse_excel import parse_excel_schedule_with_validation
        result = parse_excel_schedule_with_validation(tmp)
    finally:
        if tmp and os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except Exception:
                pass

    if not result["success"]:
        errors = result.get("errors", [])[:5]
        raise HTTPException(status_code=400, detail="; ".join(errors))

    schedule_data = result["data"]
    group = result["group"]
    conn = get_db()
    try:
        enrollment_year = user["enrollment_year"]
        if user["role"] == "assistant" or user["role"] == "admin":
            row = execute(conn,
                "SELECT enrollment_year FROM users WHERE group_name = ? LIMIT 1",
                (group,)
            ).fetchone()
            enrollment_year = row["enrollment_year"] if row else user["enrollment_year"]
        elif user["role"] == "sergeant" and group != user["group_name"]:
            conn.close()
            raise HTTPException(
                status_code=403,
                detail=f"Сержант может загружать график только своей группы. Ваша группа: {user['group_name']}"
            )

        dates = {d["date"] for d in schedule_data}
        if not dates:
            conn.close()
            return {"status": "ok", "message": "Нет записей", "count": 0}
        month_start = min(dates)
        month_ym = month_start[:7]
        from datetime import datetime as dt_klass
        try:
            dt = dt_klass.strptime(month_start, "%Y-%m-%d")
            if dt.month == 12:
                month_end_next = f"{dt.year + 1}-01-01"
            else:
                month_end_next = f"{dt.year}-{dt.month + 1:02d}-01"
        except Exception:
            month_end_next = month_start

        if overwrite != 1:
            existing = execute(conn, """
                SELECT 1 FROM duty_schedule
                WHERE group_name = ? AND enrollment_year = ?
                AND date >= ? AND date < ?
                LIMIT 1
            """, (group, enrollment_year, month_ym + "-01", month_end_next)).fetchone()
            if existing:
                conn.close()
                month_names_ru = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
                try:
                    m = int(month_ym.split("-")[1])
                    month_name = month_names_ru[m - 1]
                except Exception:
                    month_name = month_ym
                raise HTTPException(
                    status_code=409,
                    detail=f"График за {month_name} {month_ym[:4]} уже существует. Заменить?"
                )

        execute(conn,
            """DELETE FROM duty_schedule
               WHERE group_name = ? AND enrollment_year = ?
               AND date >= ? AND date < ?""",
            (group, enrollment_year, month_ym + "-01", month_end_next)
        )
        for d in schedule_data:
            execute(conn,
                """INSERT OR REPLACE INTO duty_schedule (fio, date, role, group_name, enrollment_year, gender)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (d["fio"], d["date"], d["role"], d["group"], enrollment_year, d.get("gender", "male"))
            )
        conn.commit()
        conn.close()
        month_names_ru = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
        try:
            m = int(month_ym.split("-")[1])
            month_name = month_names_ru[m - 1]
            year_str = month_ym[:4]
        except Exception:
            month_name = month_ym
            year_str = ""
        return {
            "status": "ok",
            "message": f"График за {month_name} {year_str} г. загружен. Добавлено записей: {len(schedule_data)}",
            "count": len(schedule_data),
            "month_label": f"{month_name} {year_str}",
            "month_ym": month_ym,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Сохранение графика: {e}")
        raise HTTPException(status_code=500, detail="Ошибка сохранения графика")


@app.delete("/api/schedule/month")
async def delete_schedule_month(ym: str, telegram_id: int):
    """Удалить график за месяц YYYY-MM. Сержант — только свою группу, помощник/админ — весь курс."""
    if not ym or len(ym) != 7 or ym[4] != "-":
        raise HTTPException(status_code=400, detail="Формат: YYYY-MM")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            raise HTTPException(status_code=403, detail="Нет прав")
        month_start = ym + "-01"
        try:
            y, m = int(ym[:4]), int(ym[5:7])
            if m == 12:
                month_end = f"{y + 1}-01-01"
            else:
                month_end = f"{y}-{m + 1:02d}-01"
        except Exception:
            raise HTTPException(status_code=400, detail="Неверный месяц")
        if user["role"] == "sergeant":
            execute(conn, """
                DELETE FROM duty_schedule
                WHERE enrollment_year = ? AND group_name = ? AND date >= ? AND date < ?
            """, (user["enrollment_year"], user["group_name"] or "", month_start, month_end))
        else:
            execute(conn, """
                DELETE FROM duty_schedule
                WHERE enrollment_year = ? AND date >= ? AND date < ?
            """, (user["enrollment_year"], month_start, month_end))
        conn.commit()
        conn.close()
        return {"status": "ok", "message": f"График за {ym} удалён"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] delete schedule month: {e}")
        raise HTTPException(status_code=500, detail="Ошибка удаления")


# ============================================
# 2.6. ПРАВКА ГРАФИКА: заменить/добавить наряд, контекст для форм
# ============================================

DUTY_REMOVAL_REASONS = ["заболел", "командировка", "рапорт", "другое"]


@app.get("/api/duties/edit-context")
async def get_duty_edit_context(ym: str, telegram_id: int):
    """Для правки графика за месяц: курсанты в графике и пользователи по группам (для выбора «кто заменяет»)."""
    if not ym or len(ym) != 7:
        raise HTTPException(status_code=400, detail="Формат: YYYY-MM")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав на правку графика")
        month_start = ym + "-01"
        try:
            y, m = int(ym[:4]), int(ym[5:7])
            month_end = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"
        except Exception:
            conn.close()
            raise HTTPException(status_code=400, detail="Неверный месяц")
        ey = user["enrollment_year"]
        grp = user["group_name"] or ""
        if user["role"] == "sergeant":
            schedule_rows = execute(conn, """
                SELECT DISTINCT fio, group_name FROM duty_schedule
                WHERE enrollment_year = ? AND group_name = ? AND date >= ? AND date < ?
                ORDER BY group_name, fio
            """, (ey, grp, month_start, month_end)).fetchall()
            users_rows = execute(conn, """
                SELECT fio, group_name, telegram_id FROM users
                WHERE enrollment_year = ? AND group_name = ? AND status = 'активен'
                ORDER BY fio
            """, (ey, grp)).fetchall()
        else:
            schedule_rows = execute(conn, """
                SELECT DISTINCT fio, group_name FROM duty_schedule
                WHERE enrollment_year = ? AND date >= ? AND date < ?
                ORDER BY group_name, fio
            """, (ey, month_start, month_end)).fetchall()
            users_rows = execute(conn, """
                SELECT fio, group_name, telegram_id FROM users
                WHERE enrollment_year = ? AND status = 'активен'
                ORDER BY group_name, fio
            """, (ey,)).fetchall()
        cadets_in_schedule = [{"fio": r["fio"], "group_name": r["group_name"]} for r in schedule_rows]
        group_users = [{"fio": r["fio"], "group_name": r["group_name"], "telegram_id": r["telegram_id"]} for r in users_rows]
        conn.close()
        return {"ym": ym, "cadets_in_schedule": cadets_in_schedule, "group_users": group_users, "reasons": DUTY_REMOVAL_REASONS}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] edit-context: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки контекста")


@app.get("/api/duties/role-by-fio-date")
async def get_role_by_fio_date(telegram_id: int, fio: str, date: str):
    """Для формы «Убрать наряд»: по выбранным ФИО и дате вернуть роль из графика (автоподстановка)."""
    if not telegram_id or not fio or not date or len(date) != 10:
        raise HTTPException(status_code=400, detail="Нужны telegram_id, fio, date (YYYY-MM-DD)")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав")
        ey = user["enrollment_year"]
        row = execute(conn, """
            SELECT role FROM duty_schedule
            WHERE date = ? AND enrollment_year = ? AND fio = ?
        """, (date, ey, fio.strip())).fetchone()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Нет наряда на эту дату у этого курсанта")
        return {"role": row["role"]}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] role-by-fio-date: {e}")
        raise HTTPException(status_code=500, detail="Ошибка")


@app.post("/api/duties/remove-and-replace")
async def duty_remove_and_replace(data: dict):
    """Удалить курсанта из наряда на дату и поставить вместо него другого. Запись в duty_replacements."""
    telegram_id = data.get("telegram_id")
    date = data.get("date")
    role = data.get("role")
    fio_removed = (data.get("fio_removed") or "").strip()
    fio_replacement = (data.get("fio_replacement") or "").strip()
    reason = (data.get("reason") or "заболел").strip().lower()
    if not telegram_id or not date or not role or not fio_removed or not fio_replacement:
        raise HTTPException(status_code=400, detail="Нужны: telegram_id, date, role, fio_removed, fio_replacement")
    if reason not in DUTY_REMOVAL_REASONS:
        reason = "другое"
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав")
        ey = user["enrollment_year"]
        row = execute(conn, """
            SELECT id, group_name FROM duty_schedule
            WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?
        """, (date, role, ey, fio_removed)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Запись не найдена в графике на эту дату")
        grp = row["group_name"]
        if user["role"] == "sergeant" and grp != (user["group_name"] or ""):
            conn.close()
            raise HTTPException(status_code=403, detail="Можно править только свою группу")
        execute(conn, """
            UPDATE duty_schedule SET fio = ? WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?
        """, (fio_replacement, date, role, ey, fio_removed))
        execute(conn, """
            INSERT INTO duty_replacements (date, role, group_name, enrollment_year, fio_removed, fio_replacement, reason, created_by_telegram_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (date, role, grp, ey, fio_removed, fio_replacement, reason, telegram_id))
        try:
            execute(conn,
                "UPDATE duty_shift_assignments SET fio = ? WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?",
                (fio_replacement, date, role, ey, fio_removed)
            )
            execute(conn,
                "UPDATE duty_canteen_assignments SET fio = ? WHERE date = ? AND enrollment_year = ? AND fio = ?",
                (fio_replacement, date, ey, fio_removed)
            )
        except Exception:
            pass
        conn.commit()
        editor = execute(conn,"SELECT fio FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        editor_fio = (editor["fio"] or "").split()[0] if editor else "Сержант"
        _create_schedule_notification(
            conn, grp, ey,
            f"График изменён {date}",
            f"{editor_fio}: замена {fio_removed} → {fio_replacement}. Причина: {reason}.",
            telegram_id
        )
        conn.close()
        return {"status": "ok", "message": "Замена выполнена"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] remove-and-replace: {e}")
        raise HTTPException(status_code=500, detail="Ошибка замены")


@app.post("/api/duties/add")
async def duty_add(data: dict):
    """Добавить наряд: курсант, дата, роль. Опционально — замена больному (логируем в duty_replacements)."""
    telegram_id = data.get("telegram_id")
    date = data.get("date")
    role = data.get("role")
    fio = (data.get("fio") or "").strip()
    group_name = (data.get("group_name") or "").strip()
    reason_replacing = (data.get("reason_replacing_sick") or "").strip()
    fio_replaced = (data.get("fio_replaced") or "").strip()
    if not telegram_id or not date or not role or not fio or not group_name:
        raise HTTPException(status_code=400, detail="Нужны: telegram_id, date, role, fio, group_name")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав")
        ey = user["enrollment_year"]
        if user["role"] == "sergeant" and group_name != (user["group_name"] or ""):
            conn.close()
            raise HTTPException(status_code=403, detail="Можно добавлять только в свою группу")
        gender_row = execute(conn,"SELECT gender FROM users WHERE fio = ? LIMIT 1", (fio,)).fetchone()
        gender = gender_row["gender"] if gender_row else "male"
        execute(conn, """
            INSERT OR REPLACE INTO duty_schedule (fio, date, role, group_name, enrollment_year, gender)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (fio, date, role, group_name, ey, gender))
        if reason_replacing and fio_replaced and reason_replacing.lower() in ("заболел", "1", "да", "true"):
            execute(conn, """
                INSERT INTO duty_replacements (date, role, group_name, enrollment_year, fio_removed, fio_replacement, reason, created_by_telegram_id)
                VALUES (?, ?, ?, ?, ?, ?, 'заболел', ?)
            """, (date, role, group_name, ey, fio_replaced, fio, telegram_id))
        conn.commit()
        editor = execute(conn,"SELECT fio FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        editor_fio = (editor["fio"] or "").split()[0] if editor else "Сержант"
        _create_schedule_notification(
            conn, group_name, ey,
            f"График изменён {date}",
            f"{editor_fio}: добавлен наряд для {fio} ({role}).",
            telegram_id
        )
        conn.close()
        return {"status": "ok", "message": "Наряд добавлен"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] duty add: {e}")
        raise HTTPException(status_code=500, detail="Ошибка добавления")


@app.post("/api/duties/remove")
async def duty_remove(data: dict):
    """Удалить наряд за дату у курсанта (без замены)."""
    telegram_id = data.get("telegram_id")
    date = data.get("date")
    fio_removed = (data.get("fio_removed") or "").strip()
    role = (data.get("role") or "").strip()
    if not telegram_id or not date or not fio_removed or not role:
        raise HTTPException(status_code=400, detail="Нужны: telegram_id, date, fio_removed, role")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав")
        ey = user["enrollment_year"]
        row = execute(conn, """
            SELECT group_name FROM duty_schedule
            WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?
        """, (date, role, ey, fio_removed)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Запись не найдена в графике на эту дату")
        if user["role"] == "sergeant" and row["group_name"] != (user["group_name"] or ""):
            conn.close()
            raise HTTPException(status_code=403, detail="Можно править только свою группу")
        execute(conn, """
            DELETE FROM duty_schedule
            WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?
        """, (date, role, ey, fio_removed))
        try:
            execute(conn,"DELETE FROM duty_shift_assignments WHERE date = ? AND role = ? AND enrollment_year = ? AND fio = ?",
                         (date, role, ey, fio_removed))
            execute(conn,"DELETE FROM duty_canteen_assignments WHERE date = ? AND enrollment_year = ? AND fio = ?",
                         (date, ey, fio_removed))
        except Exception:
            pass
        conn.commit()
        editor = execute(conn,"SELECT fio FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        editor_fio = (editor["fio"] or "").split()[0] if editor else "Сержант"
        _create_schedule_notification(
            conn, row["group_name"], ey,
            f"График изменён {date}",
            f"{editor_fio}: снят с наряда {fio_removed} (роль: {role}).",
            telegram_id
        )
        conn.close()
        return {"status": "ok", "message": "Наряд удалён"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] duty remove: {e}")
        raise HTTPException(status_code=500, detail="Ошибка удаления")


@app.post("/api/duties/change-role")
async def duty_change_role(data: dict):
    """Изменить роль курсанта на дату (например: был Курс — стал ГБР)."""
    telegram_id = data.get("telegram_id")
    date = data.get("date")
    fio = (data.get("fio") or "").strip()
    new_role = (data.get("new_role") or "").strip().lower()
    if not telegram_id or not date or not fio or not new_role:
        raise HTTPException(status_code=400, detail="Нужны: telegram_id, date, fio, new_role")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            conn.close()
            raise HTTPException(status_code=403, detail="Нет прав")
        ey = user["enrollment_year"]
        row = execute(conn, """
            SELECT role, group_name FROM duty_schedule
            WHERE date = ? AND enrollment_year = ? AND fio = ?
        """, (date, ey, fio)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Наряд на эту дату не найден")
        if user["role"] == "sergeant" and row["group_name"] != (user["group_name"] or ""):
            conn.close()
            raise HTTPException(status_code=403, detail="Можно править только свою группу")
        execute(conn, """
            UPDATE duty_schedule SET role = ? WHERE date = ? AND enrollment_year = ? AND fio = ?
        """, (new_role, date, ey, fio))
        try:
            execute(conn,"DELETE FROM duty_shift_assignments WHERE date = ? AND enrollment_year = ? AND fio = ?", (date, ey, fio))
            execute(conn,"DELETE FROM duty_canteen_assignments WHERE date = ? AND enrollment_year = ? AND fio = ?", (date, ey, fio))
        except Exception:
            pass
        conn.commit()
        editor = execute(conn,"SELECT fio FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        editor_fio = (editor["fio"] or "").split()[0] if editor else "Сержант"
        _create_schedule_notification(
            conn, row["group_name"], ey,
            f"График изменён {date}",
            f"{editor_fio}: у {fio} изменена роль на {new_role}.",
            telegram_id
        )
        conn.close()
        return {"status": "ok", "message": "Роль изменена"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] duty change-role: {e}")
        raise HTTPException(status_code=500, detail="Ошибка изменения")


# ============================================
# 2.7. УВЕДОМЛЕНИЯ
# ============================================

@app.get("/api/notifications")
async def get_notifications(telegram_id: int, limit: int = 50):
    """Список уведомлений для пользователя: свои + по группе + по курсу + общие. С флагом прочитано."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user:
            conn.close()
            return {"items": [], "unread_count": 0}
        grp, ey = (user["group_name"] or ""), user["enrollment_year"]
        rows = execute(conn, """
            SELECT n.id, n.title, n.body, n.type, n.created_at, n.scope, n.scope_value,
                   r.telegram_id IS NOT NULL AS read
            FROM notifications n
            LEFT JOIN notification_read r ON r.notification_id = n.id AND r.telegram_id = ?
            WHERE (n.telegram_id = ? OR (n.telegram_id IS NULL AND (
                (n.scope = 'group' AND n.scope_value = ?) OR
                (n.scope = 'course' AND n.scope_value = ?) OR
                n.scope = 'all'
            )))
            ORDER BY n.created_at DESC
            LIMIT ?
        """, (telegram_id, telegram_id, grp, str(ey), limit)).fetchall()
        items = []
        unread = 0
        for r in rows:
            read = bool(r["read"])
            if not read:
                unread += 1
            items.append({
                "id": r["id"],
                "title": r["title"],
                "body": r["body"] or "",
                "type": r["type"] or "info",
                "created_at": r["created_at"],
                "read": read,
            })
        conn.close()
        return {"items": items, "unread_count": unread}
    except Exception as e:
        print(f"[ERROR] notifications: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки уведомлений")


@app.post("/api/notifications/read")
async def mark_notifications_read(data: dict):
    """Отметить уведомления как прочитанные. notification_ids: список id или "all"."""
    telegram_id = data.get("telegram_id")
    notification_ids = data.get("notification_ids")
    if not telegram_id:
        raise HTTPException(status_code=400, detail="Нужен telegram_id")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        if notification_ids == "all" or (isinstance(notification_ids, list) and len(notification_ids) == 0):
            # Получить все id уведомлений, которые пользователь видит и не прочитал
            user = execute(conn,"SELECT group_name, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
            if not user:
                conn.close()
                return {"status": "ok", "marked": 0}
            grp, ey = (user["group_name"] or ""), user["enrollment_year"]
            ids = [row["id"] for row in execute(conn, """
                SELECT n.id FROM notifications n
                LEFT JOIN notification_read r ON r.notification_id = n.id AND r.telegram_id = ?
                WHERE (n.telegram_id = ? OR (n.scope = 'group' AND n.scope_value = ?) OR (n.scope = 'course' AND n.scope_value = ?) OR n.scope = 'all')
                  AND r.telegram_id IS NULL
            """, (telegram_id, telegram_id, grp, str(ey))).fetchall()]
        else:
            ids = [int(x) for x in notification_ids] if isinstance(notification_ids, list) else []
        for nid in ids:
            execute(conn,"INSERT OR IGNORE INTO notification_read (notification_id, telegram_id) VALUES (?, ?)", (nid, telegram_id))
        conn.commit()
        conn.close()
        return {"status": "ok", "marked": len(ids)}
    except Exception as e:
        print(f"[ERROR] notifications read: {e}")
        raise HTTPException(status_code=500, detail="Ошибка")


# ============================================
# 2.8. РЕЙТИНГ (очки из нарядов по весам опроса)
# ============================================

def _get_user_duty_points(conn, telegram_id: int, month_from: str = None, month_to: str = None):
    """Сумма баллов пользователя за наряды (по duty_schedule + object_weights). Период опционально."""
    user = execute(conn,"SELECT fio, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
    if not user:
        return 0.0
    fio, ey = user["fio"], user["enrollment_year"]
    extra = ""
    params = [fio, ey]
    if month_from:
        extra += " AND date >= ?"
        params.append(month_from + "-01")
    if month_to:
        y, m = int(month_to[:4]), int(month_to[5:7])
        if m == 12:
            month_end = f"{y+1}-01-01"
        else:
            month_end = f"{y}-{m+1:02d}-01"
        extra += " AND date < ?"
        params.append(month_end)
    rows = execute(conn,f"""
        SELECT ds.role, ds.date FROM duty_schedule ds
        WHERE ds.fio = ? AND ds.enrollment_year = ? {extra}
    """, params).fetchall()
    # Веса: duty_objects (name = Курс, ГБР, Столовая, ЗУБ и дочерние для столовой) + object_weights
    weights = {}
    for w in execute(conn,"SELECT o.name, o.parent_id, ow.weight FROM duty_objects o JOIN object_weights ow ON ow.object_id = o.id").fetchall():
        key = (w["name"], w["parent_id"])
        weights[key] = float(w["weight"] or 0)
    # Роль в графике может быть кодом (к, гбр, с ...). Соответствие с duty_objects по имени
    role_to_name = {"к": "Курс", "дк": "Дежурный по курсу", "с": "Столовая", "гбр": "ГБР", "зуб": "ЗУБ", "путсо": "ПУТСО", "м": "Медчасть"}
    total = 0.0
    for r in rows:
        role_code = (r["role"] or "").lower()
        name = role_to_name.get(role_code, role_code)
        # Ищем вес по имени (parent_id NULL для основных)
        for (obj_name, parent_id), w in weights.items():
            if obj_name == name and parent_id is None:
                total += w
                break
        else:
            total += 10.0  # дефолт если нет веса
    return round(total, 1)


@app.get("/api/rating/me")
async def rating_me(telegram_id: int):
    """Очки пользователя, место в топе по курсу и по институту (за всё время)."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT enrollment_year, group_name FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user:
            conn.close()
            return {"points": 0, "rank_course": None, "rank_institute": None}
        ey = user["enrollment_year"]
        points = _get_user_duty_points(conn, telegram_id, None, None)
        # Ранг по курсу и институту — считаем очки для всех и ранжируем
        all_course = []
        for row in execute(conn,"SELECT telegram_id FROM users WHERE enrollment_year = ? AND status = 'активен'", (ey,)).fetchall():
            tid = row["telegram_id"]
            p = _get_user_duty_points(conn, tid, None, None)
            all_course.append((tid, p))
        all_course.sort(key=lambda x: -x[1])
        rank_course = next((i + 1 for i, (tid, _) in enumerate(all_course) if tid == telegram_id), None)
        all_inst = []
        for row in execute(conn,"SELECT telegram_id FROM users WHERE status = 'активен'").fetchall():
            p = _get_user_duty_points(conn, row["telegram_id"], None, None)
            all_inst.append((row["telegram_id"], p))
        all_inst.sort(key=lambda x: -x[1])
        rank_institute = next((i + 1 for i, (tid, _) in enumerate(all_inst) if tid == telegram_id), None)
        conn.close()
        return {"points": points, "rank_course": rank_course, "rank_institute": rank_institute}
    except Exception as e:
        print(f"[ERROR] rating/me: {e}")
        raise HTTPException(status_code=500, detail="Ошибка рейтинга")


@app.get("/api/rating/top")
async def rating_top(telegram_id: int, period: str = "all", scope: str = "course", limit: int = 30):
    """Топ по очкам. period: month | all, scope: course | institute."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,"SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            conn.close()
            return {"top": []}
        ey = user["enrollment_year"]
        month_from = datetime.now().strftime("%Y-%m") if period == "month" else None
        month_to = datetime.now().strftime("%Y-%m") if period == "month" else None
        if scope == "course":
            rows = execute(conn,"SELECT telegram_id, fio, group_name FROM users WHERE enrollment_year = ? AND status = 'активен'", (ey,)).fetchall()
        else:
            rows = execute(conn,"SELECT telegram_id, fio, group_name FROM users WHERE status = 'активен'").fetchall()
        result = []
        for r in rows:
            p = _get_user_duty_points(conn, r["telegram_id"], month_from, month_to)
            result.append({"telegram_id": r["telegram_id"], "fio": r["fio"], "group_name": r["group_name"], "points": p})
        result.sort(key=lambda x: -x["points"])
        result = result[:limit]
        for i, row in enumerate(result):
            row["rank"] = i + 1
        conn.close()
        return {"top": result, "period": period, "scope": scope}
    except Exception as e:
        print(f"[ERROR] rating/top: {e}")
        raise HTTPException(status_code=500, detail="Ошибка рейтинга")


# ============================================
# 2.9. ДОСТИЖЕНИЯ
# ============================================

def _unlock_achievements(conn, telegram_id: int):
    """Проверяет условия достижений и добавляет в user_achievements при выполнении."""
    try:
        user = execute(conn,"SELECT fio, enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            return
        fio, ey = user["fio"], user["enrollment_year"]
        # first_10_duties
        count_duties = execute(conn,"SELECT COUNT(*) FROM duty_schedule WHERE fio = ? AND enrollment_year = ?", (fio, ey)).fetchone()[0]
        if count_duties >= 10:
            execute(conn,"INSERT OR IGNORE INTO user_achievements (telegram_id, achievement_id) VALUES (?, 'first_10_duties')", (telegram_id,))
        # top3_course, top10_institute — по текущим очкам
        points = _get_user_duty_points(conn, telegram_id, None, None)
        course_list = []
        for row in execute(conn,"SELECT telegram_id FROM users WHERE enrollment_year = ? AND status = 'активен'", (ey,)).fetchall():
            p = _get_user_duty_points(conn, row["telegram_id"], None, None)
            course_list.append((row["telegram_id"], p))
        course_list.sort(key=lambda x: -x[1])
        rank_course = next((i + 1 for i, (tid, _) in enumerate(course_list) if tid == telegram_id), None)
        if rank_course is not None and rank_course <= 3:
            execute(conn,"INSERT OR IGNORE INTO user_achievements (telegram_id, achievement_id) VALUES (?, 'top3_course')", (telegram_id,))
        all_inst = []
        for row in execute(conn,"SELECT telegram_id FROM users WHERE status = 'активен'").fetchall():
            p = _get_user_duty_points(conn, row["telegram_id"], None, None)
            all_inst.append((row["telegram_id"], p))
        all_inst.sort(key=lambda x: -x[1])
        rank_inst = next((i + 1 for i, (tid, _) in enumerate(all_inst) if tid == telegram_id), None)
        if rank_inst is not None and rank_inst <= 10:
            execute(conn,"INSERT OR IGNORE INTO user_achievements (telegram_id, achievement_id) VALUES (?, 'top10_institute')", (telegram_id,))
        conn.commit()
    except Exception as e:
        print(f"[WARN] _unlock_achievements: {e}")


@app.get("/api/achievements")
async def get_achievements(telegram_id: int):
    """Список всех достижений с флагом получено и процентом обладателей."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        _unlock_achievements(conn, telegram_id)
        total_users = execute(conn,"SELECT COUNT(*) FROM users WHERE status = 'активен'").fetchone()[0] or 1
        rows = execute(conn, """
            SELECT a.id, a.title, a.description, a.icon_url, a.sort_order,
                   ua.telegram_id IS NOT NULL AS unlocked
            FROM achievements a
            LEFT JOIN user_achievements ua ON ua.achievement_id = a.id AND ua.telegram_id = ?
            ORDER BY a.sort_order, a.id
        """, (telegram_id,)).fetchall()
        result = []
        for r in rows:
            count = execute(conn,"SELECT COUNT(*) FROM user_achievements WHERE achievement_id = ?", (r["id"],)).fetchone()[0]
            pct = round(100.0 * count / total_users, 1) if total_users else 0
            result.append({
                "id": r["id"],
                "title": r["title"],
                "description": r["description"] or "",
                "icon_url": r["icon_url"],
                "unlocked": bool(r["unlocked"]),
                "percent_owners": pct,
            })
        conn.close()
        return {"achievements": result}
    except Exception as e:
        print(f"[ERROR] achievements: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки достижений")


# ============================================
# 5. ОПРОСНИК (SURVEY) API — попарное сравнение 2/1/0
# ============================================

def _get_all_pairs(objects):
    """Генерирует все пары из списка объектов [(id, name), ...]"""
    pairs = []
    for i in range(len(objects)):
        for j in range(i + 1, len(objects)):
            pairs.append({
                "object_a": {"id": objects[i]["id"], "name": objects[i]["name"]},
                "object_b": {"id": objects[j]["id"], "name": objects[j]["name"]}
            })
    return pairs


@app.get("/api/survey/list")
async def get_survey_list(telegram_id: int):
    """Список опросов: системные (юноши/девушки) и пользовательские для группы/курса."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT gender, group_name, enrollment_year, role FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
        if not user:
            conn.close()
            return {"system": [], "custom": []}
        gender = user["gender"] or "male"
        group_name = user["group_name"] or ""
        enrollment_year = user["enrollment_year"]
        role = user["role"] or "user"

        system = [
            {"id": "male", "title": "Опрос для парней (сложность нарядов)", "for_gender": "male"},
            {"id": "female", "title": "Опрос для девушек (ПУТСО, Столовая, Медчасть)", "for_gender": "female"},
        ]

        custom_rows = execute(conn, """
            SELECT s.id, s.title, s.scope_type, s.scope_value, s.created_by_telegram_id, s.ends_at, s.completed_at
            FROM custom_surveys s
            WHERE s.completed_at IS NULL
            AND (
                (s.scope_type = 'group' AND s.scope_value = ?)
                OR (s.scope_type = 'course' AND s.scope_value = ?)
                OR (s.scope_type = 'system')
            )
            ORDER BY s.created_at DESC
        """, (group_name, str(enrollment_year))).fetchall()
        custom = []
        for r in custom_rows:
            custom.append({
                "id": r["id"],
                "title": r["title"],
                "scope_type": r["scope_type"],
                "scope_value": r["scope_value"],
                "created_by_telegram_id": r["created_by_telegram_id"],
                "ends_at": r["ends_at"],
                "completed_at": r["completed_at"],
                "can_complete": r["created_by_telegram_id"] == telegram_id or role in ("admin", "assistant"),
            })
        conn.close()
        return {"system": system, "custom": custom, "user_gender": gender}
    except Exception as e:
        print(f"[ERROR] Ошибка списка опросов: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.get("/api/survey/pairs")
async def get_survey_pairs(stage: str = "main"):
    """
    Возвращает пары для попарного голосования.
    stage: 'main' — 4 основных наряда (6 пар), 'canteen' — 14 случайных пар из 11 объектов столовой.
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        if stage == "main":
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id IS NULL AND name != 'Опрос девушек' ORDER BY id"
            ).fetchall()
        elif stage == "female":
            female_parent = execute(conn,
                "SELECT id FROM duty_objects WHERE name = 'Опрос девушек' AND parent_id IS NULL"
            ).fetchone()
            if not female_parent:
                conn.close()
                return {"pairs": [], "stage": stage}
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id = ? ORDER BY id",
                (female_parent["id"],)
            ).fetchall()
            objects = [{"id": r["id"], "name": r["name"]} for r in rows]
            pairs = _get_all_pairs(objects)
            conn.close()
            return {"pairs": pairs, "stage": stage}
        else:  # canteen — 6 объектов столовой, все возможные пары без повторений (15 пар)
            CANTEEN_OBJECT_NAMES = [
                "Горячий цех", "Овощной цех", "Стаканы", "Железо", "Лента", "Тарелки"
            ]
            canteen = execute(conn,
                "SELECT id FROM duty_objects WHERE name='Столовая' AND parent_id IS NULL"
            ).fetchone()
            if not canteen:
                conn.close()
                return {"pairs": [], "stage": stage}
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id = ? ORDER BY id",
                (canteen["id"],)
            ).fetchall()
            by_name = {r["name"]: r for r in rows}
            objects = []
            for display_name in CANTEEN_OBJECT_NAMES:
                r = by_name.get(display_name) or (by_name.get("Мойка-тарелки") if display_name == "Тарелки" else None)
                if r:
                    objects.append({"id": r["id"], "name": display_name})
            seen = set()
            objects = [o for o in objects if o["id"] not in seen and not seen.add(o["id"])]
            pairs = _get_all_pairs(objects)
            random.shuffle(pairs)
            conn.close()
            return {"pairs": pairs, "stage": stage}
        
        objects = [{"id": r["id"], "name": r["name"]} for r in rows]
        pairs = _get_all_pairs(objects)
        random.shuffle(pairs)
        conn.close()
        return {"pairs": pairs, "stage": stage}
    except Exception as e:
        print(f"[ERROR] Ошибка получения пар: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.post("/api/survey/pair-vote")
async def submit_pair_vote(data: dict):
    """
    Принимает голос попарного сравнения.
    choice: 'a' — первый сложнее (2/0), 'b' — второй сложнее (0/2), 'equal' — одинаково (1/1)
    """
    user_id = data.get('user_id')
    object_a_id = data.get('object_a_id')
    object_b_id = data.get('object_b_id')
    choice = data.get('choice')
    stage = data.get('stage', 'main')
    if stage not in ('main', 'canteen', 'female'):
        stage = 'main'

    if not all([user_id, object_a_id, object_b_id, choice]) or choice not in ('a', 'b', 'equal'):
        raise HTTPException(status_code=400, detail="Неверные данные")

    # Упорядочиваем пару: object_a_id < object_b_id для уникальности
    oa, ob = int(object_a_id), int(object_b_id)
    if oa > ob:
        oa, ob = ob, oa
        choice = 'b' if choice == 'a' else ('a' if choice == 'b' else 'equal')

    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")

    try:
        user = execute(conn,"SELECT id FROM users WHERE telegram_id = ?", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        db_user_id = user['id']

        execute(conn, """
            INSERT INTO survey_pair_votes (user_id, object_a_id, object_b_id, choice, stage)
            VALUES (?, ?, ?, ?, ?)
        """, (db_user_id, oa, ob, choice, stage))
        conn.commit()

        # Количество уникальных проголосовавших
        voted_count = execute(conn,
            "SELECT COUNT(DISTINCT user_id) as cnt FROM survey_pair_votes"
        ).fetchone()['cnt']

        # При 100 голосах можно автоматически финализировать (вызывать расчёт весов)
        # Пока оставляем ручную финализацию через админа
        conn.close()
        return {"status": "ok", "message": "Голос учтён", "total_voted": voted_count}
    except DBIntegrityError:
        raise HTTPException(status_code=409, detail="Вы уже голосовали за эту пару")
    except Exception as e:
        print(f"[ERROR] Ошибка голосования: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")
    finally:
        if conn:
            conn.close()


@app.get("/api/survey/status")
async def get_survey_status():
    """Возвращает статистику опроса: сколько проголосовало из скольких"""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        total_users = execute(conn,"SELECT COUNT(*) as cnt FROM users WHERE status='активен'").fetchone()['cnt']
        try:
            voted_users = execute(conn,
                "SELECT COUNT(DISTINCT user_id) as cnt FROM survey_pair_votes"
            ).fetchone()['cnt']
        except Exception:
            voted_users = 0
        return {"total": total_users, "voted": voted_users}
    except Exception as e:
        print(f"[ERROR] Ошибка статуса опроса: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")
    finally:
        conn.close()


def _clamp_coef(k: float, max_k: float = 2.0) -> float:
    """Нижнее значение 0.8, верхний — max_k (основные наряды 2.0, столовые 1.5)."""
    return max(0.8, min(max_k, k))


def _calc_weights_from_pair_votes(conn, stage_filter: str = None):
    """
    Рассчитывает веса по формуле: S = сумма баллов объекта, avg = среднее, k = S/avg, вес = 10 × k.
    stage_filter: None = все этапы, 'main' | 'canteen' | 'female' = только этот этап.
    """
        # Этап 1: основные наряды
    if stage_filter is None or stage_filter == 'main':
        main_ids = [r['id'] for r in execute(conn,
            "SELECT id FROM duty_objects WHERE parent_id IS NULL ORDER BY id"
        ).fetchall()]

        # Считаем баллы: choice 'a' → object_a +2, object_b +0; 'b' → object_a +0, object_b +2; 'equal' → +1 каждому
        scores = {oid: 0.0 for oid in main_ids}
        votes = execute(conn,
            "SELECT object_a_id, object_b_id, choice FROM survey_pair_votes WHERE stage='main'"
        ).fetchall()
        for v in votes:
            a, b = v['object_a_id'], v['object_b_id']
            if v['choice'] == 'a':
                scores[a] = scores.get(a, 0) + 2
                scores[b] = scores.get(b, 0) + 0
            elif v['choice'] == 'b':
                scores[a] = scores.get(a, 0) + 0
                scores[b] = scores.get(b, 0) + 2
            else:
                scores[a] = scores.get(a, 0) + 1
                scores[b] = scores.get(b, 0) + 1

        if len(main_ids) > 0:
            total = sum(scores.get(i, 0) for i in main_ids)
            avg = total / len(main_ids) if total > 0 else 1
            for oid in main_ids:
                s = scores.get(oid, 0)
                k = (s / avg) if (avg > 0 and s > 0) else 0.8
                k = _clamp_coef(k)
                w = 10 * k
                execute(conn, """
                    INSERT INTO object_weights (object_id, weight) VALUES (?, ?)
                    ON CONFLICT(object_id) DO UPDATE SET weight=excluded.weight, calculated_at=CURRENT_TIMESTAMP
                """, (oid, w))

    # Этап 2: объекты столовой — веса относительные к весу столовой
    if stage_filter is None or stage_filter == 'canteen':
        canteen_row = execute(conn,
            "SELECT id FROM duty_objects WHERE name='Столовая' AND parent_id IS NULL"
        ).fetchone()
        if canteen_row:
            canteen_id = canteen_row['id']
            canteen_weight_row = execute(conn,
                "SELECT weight FROM object_weights WHERE object_id = ?", (canteen_id,)
            ).fetchone()
            canteen_weight = canteen_weight_row['weight'] if canteen_weight_row else 10

            sub_ids = [r['id'] for r in execute(conn,
                "SELECT id FROM duty_objects WHERE parent_id = ? ORDER BY id", (canteen_id,)
            ).fetchall()]

            sub_scores = {oid: 0.0 for oid in sub_ids}
            sub_votes = execute(conn,
                "SELECT object_a_id, object_b_id, choice FROM survey_pair_votes WHERE stage='canteen'"
            ).fetchall()
            for v in sub_votes:
                a, b = v['object_a_id'], v['object_b_id']
                if v['choice'] == 'a':
                    sub_scores[a] = sub_scores.get(a, 0) + 2
                    sub_scores[b] = sub_scores.get(b, 0) + 0
                elif v['choice'] == 'b':
                    sub_scores[a] = sub_scores.get(a, 0) + 0
                    sub_scores[b] = sub_scores.get(b, 0) + 2
                else:
                    sub_scores[a] = sub_scores.get(a, 0) + 1
                    sub_scores[b] = sub_scores.get(b, 0) + 1

            if len(sub_ids) > 0:
                sub_total = sum(sub_scores.get(i, 0) for i in sub_ids)
                sub_avg = sub_total / len(sub_ids) if sub_total > 0 else 1
                for oid in sub_ids:
                    s = sub_scores.get(oid, 0)
                    k_sub = (s / sub_avg) if (sub_avg > 0 and s > 0) else 0.8
                    k_sub = _clamp_coef(k_sub, max_k=1.6)
                    w_sub = canteen_weight * k_sub
                    execute(conn, """
                        INSERT INTO object_weights (object_id, weight) VALUES (?, ?)
                        ON CONFLICT(object_id) DO UPDATE SET weight=excluded.weight, calculated_at=CURRENT_TIMESTAMP
                    """, (oid, w_sub))

    # Этап 3: опрос для девушек (ПУТСО, Столовая, Медчасть)
    if stage_filter is None or stage_filter == 'female':
        female_parent = execute(conn,
            "SELECT id FROM duty_objects WHERE name = 'Опрос девушек' AND parent_id IS NULL"
        ).fetchone()
        if female_parent:
            female_ids = [r["id"] for r in execute(conn,
                "SELECT id FROM duty_objects WHERE parent_id = ? ORDER BY id", (female_parent["id"],)
            ).fetchall()]
            if female_ids:
                f_scores = {oid: 0.0 for oid in female_ids}
                f_votes = execute(conn,
                    "SELECT object_a_id, object_b_id, choice FROM survey_pair_votes WHERE stage = 'female'"
                ).fetchall()
                for v in f_votes:
                    a, b = v["object_a_id"], v["object_b_id"]
                    if v["choice"] == "a":
                        f_scores[a] = f_scores.get(a, 0) + 2
                        f_scores[b] = f_scores.get(b, 0) + 0
                    elif v["choice"] == "b":
                        f_scores[a] = f_scores.get(a, 0) + 0
                        f_scores[b] = f_scores.get(b, 0) + 2
                    else:
                        f_scores[a] = f_scores.get(a, 0) + 1
                        f_scores[b] = f_scores.get(b, 0) + 1
                f_total = sum(f_scores.get(i, 0) for i in female_ids)
                f_avg = f_total / len(female_ids) if f_total > 0 else 1
                for oid in female_ids:
                    s = f_scores.get(oid, 0)
                    k = (s / f_avg) if (f_avg > 0 and s > 0) else 0.8
                    k = _clamp_coef(k, max_k=1.6)
                    w = 10 * k
                    execute(conn, """
                        INSERT INTO object_weights (object_id, weight) VALUES (?, ?)
                        ON CONFLICT(object_id) DO UPDATE SET weight=excluded.weight, calculated_at=CURRENT_TIMESTAMP
                    """, (oid, w))


@app.post("/api/survey/finalize")
async def finalize_survey(data: dict):
    """Завершает опрос и вычисляет веса по формуле k = S/avg, итог = 10 × k"""
    admin_id = data.get('admin_id')
    if not admin_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,"SELECT role FROM users WHERE telegram_id = ?", (admin_id,)).fetchone()
        if not user or user['role'] not in ('admin', 'assistant'):
            raise HTTPException(status_code=403, detail="Недостаточно прав")
    except Exception:
        raise HTTPException(status_code=500, detail="Ошибка проверки прав")
    finally:
        conn.close()

    conn = get_db()
    try:
        stage = (data.get('stage') or '').strip() or None
        if stage and stage not in ('main', 'canteen', 'female'):
            stage = None
        voted = execute(conn,"SELECT COUNT(DISTINCT user_id) as cnt FROM survey_pair_votes").fetchone()["cnt"]
        _calc_weights_from_pair_votes(conn, stage_filter=stage)
        conn.commit()
        from datetime import date as date_type
        today = date_type.today()
        next_month = today.month + 1 if today.month < 12 else 1
        next_year = today.year if today.month < 12 else today.year + 1
        next_label = f"{next_year}-{next_month:02d}"
        return {"status": "ok", "message": "Веса вычислены и сохранены", "total_voted": voted, "stage": stage, "next_period": next_label}
    except Exception as e:
        print(f"[ERROR] Ошибка финализации опроса: {e}")
        raise HTTPException(status_code=500, detail="Ошибка вычисления")
    finally:
        conn.close()


@app.get("/api/survey/pair-stats")
async def get_survey_pair_stats(stage: str = "main"):
    """Для визуализации: по каждой паре — число ответов A сложнее / равно / B сложнее и доли в %."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        if stage == "female":
            female_parent = execute(conn,
                "SELECT id FROM duty_objects WHERE name = 'Опрос девушек' AND parent_id IS NULL"
            ).fetchone()
            if not female_parent:
                conn.close()
                return {"pairs": []}
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id = ? ORDER BY id",
                (female_parent["id"],)
            ).fetchall()
        elif stage == "canteen":
            canteen = execute(conn,
                "SELECT id FROM duty_objects WHERE name = 'Столовая' AND parent_id IS NULL"
            ).fetchone()
            if not canteen:
                conn.close()
                return {"pairs": []}
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id = ? ORDER BY id",
                (canteen["id"],)
            ).fetchall()
        else:
            rows = execute(conn,
                "SELECT id, name FROM duty_objects WHERE parent_id IS NULL AND name != 'Опрос девушек' ORDER BY id"
            ).fetchall()
        id2name = {r["id"]: r["name"] for r in rows}
        votes = execute(conn,
            "SELECT object_a_id, object_b_id, choice FROM survey_pair_votes WHERE stage = ?",
            (stage,)
        ).fetchall()
        from collections import defaultdict
        pair_counts = defaultdict(lambda: {"a": 0, "b": 0, "equal": 0})
        for v in votes:
            a, b = v["object_a_id"], v["object_b_id"]
            key = (min(a, b), max(a, b))
            pair_counts[key][v["choice"]] += 1
        pairs = []
        for (oa, ob), counts in pair_counts.items():
            total = counts["a"] + counts["b"] + counts["equal"]
            if total == 0:
                continue
            name_a = id2name.get(oa, "?")
            name_b = id2name.get(ob, "?")
            pairs.append({
                "object_a_name": name_a,
                "object_b_name": name_b,
                "count_a": counts["a"],
                "count_b": counts["b"],
                "count_equal": counts["equal"],
                "total": total,
                "pct_a": round(100 * counts["a"] / total, 1),
                "pct_b": round(100 * counts["b"] / total, 1),
                "pct_equal": round(100 * counts["equal"] / total, 1),
            })
        conn.close()
        return {"pairs": pairs, "stage": stage}
    except Exception as e:
        print(f"[ERROR] pair-stats: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.get("/api/survey/results")
async def get_survey_results():
    """Возвращает вычисленные веса объектов для отображения"""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        # Возвращаем все объекты с их весами (если есть)
        # SQLite не поддерживает NULLS FIRST, поэтому сортируем так:
        # сначала объекты без родителя (parent_id IS NULL), потом с родителем
        results = execute(conn, """
            SELECT o.id, o.name, o.parent_id, w.weight
            FROM duty_objects o
            LEFT JOIN object_weights w ON o.id = w.object_id
            ORDER BY o.parent_id IS NULL DESC, o.parent_id, o.name
        """).fetchall()
        return [dict(r) for r in results]
    except Exception as e:
        print(f"[ERROR] Ошибка получения результатов: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")
    finally:
        conn.close()


@app.get("/api/survey/user-results")
async def get_user_survey_results(telegram_id: int):
    """Возвращает результаты опроса для конкретного пользователя (если он прошёл опрос)"""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        # Проверяем, проходил ли пользователь опрос
        user = execute(conn,
            "SELECT id, gender FROM users WHERE telegram_id = ?", (telegram_id,)
        ).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        db_user_id = user["id"]
        user_gender = (user["gender"] or "male").strip().lower()

        # Для девушек — считаем пройденным опрос stage=female; для юношей — main/canteen
        if user_gender == "female":
            voted = execute(conn,
                "SELECT 1 FROM survey_pair_votes WHERE user_id = ? AND stage = 'female' LIMIT 1",
                (db_user_id,)
            ).fetchone() is not None
            female_parent = execute(conn,
                "SELECT id FROM duty_objects WHERE name = 'Опрос девушек' AND parent_id IS NULL"
            ).fetchone()
            if voted and female_parent:
                results = execute(conn, """
                    SELECT o.id, o.name, o.parent_id, w.weight as median_weight
                    FROM duty_objects o
                    LEFT JOIN object_weights w ON o.id = w.object_id
                    WHERE o.parent_id = ?
                    ORDER BY o.name
                """, (female_parent["id"],)).fetchall()
                conn.close()
                return {"voted": True, "results": [dict(r) for r in results], "survey_stage": "female"}
        else:
            voted_main = execute(conn,
                "SELECT 1 FROM survey_pair_votes WHERE user_id = ? AND stage = 'main' LIMIT 1",
                (db_user_id,)
            ).fetchone() is not None
            voted_canteen = execute(conn,
                "SELECT 1 FROM survey_pair_votes WHERE user_id = ? AND stage = 'canteen' LIMIT 1",
                (db_user_id,)
            ).fetchone() is not None
            voted = voted_main or voted_canteen
        if not voted:
            conn.close()
            return {"voted": False, "message": "Вы ещё не прошли опрос"}

        # Юноши: возвращаем по этапам и результаты (веса основных нарядов и столовой)
        female_parent = execute(conn,
            "SELECT id FROM duty_objects WHERE name = 'Опрос девушек' AND parent_id IS NULL"
        ).fetchone()
        female_id = female_parent["id"] if female_parent else -1
        results = execute(conn, """
            SELECT o.id, o.name, o.parent_id, w.weight as median_weight
            FROM duty_objects o
            LEFT JOIN object_weights w ON o.id = w.object_id
            WHERE o.id != ? AND (o.parent_id IS NULL OR o.parent_id != ?)
            ORDER BY (o.parent_id IS NULL) DESC, o.parent_id, o.name
        """, (female_id, female_id)).fetchall()

        return {
            "voted": True,
            "voted_main": voted_main,
            "voted_canteen": voted_canteen,
            "results": [dict(r) for r in results],
            "survey_stage": "main" if voted_main else "canteen"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Ошибка получения результатов пользователя: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")
    finally:
        conn.close()


# ============================================
# 5b. ПОЛЬЗОВАТЕЛЬСКИЕ ОПРОСЫ (custom)
# ============================================
@app.post("/api/survey/custom")
async def create_custom_survey(data: dict):
    """Создать опрос: сержант — группа, помощник — курс, админ — группа/курс/системный."""
    telegram_id = data.get("telegram_id")
    title = (data.get("title") or "").strip()
    scope_type = data.get("scope_type")  # 'group' | 'course' | 'system'
    options = data.get("options") or []  # ["Вариант 1", "Вариант 2", ...]
    ends_at = data.get("ends_at")  # optional ISO date/datetime
    if not telegram_id or not title or scope_type not in ("group", "course", "system") or len(options) < 2:
        raise HTTPException(status_code=400, detail="Нужны: telegram_id, title, scope_type (group|course|system), options (минимум 2)")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn,
            "SELECT role, group_name, enrollment_year FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
        if not user or user["role"] not in ("sergeant", "assistant", "admin"):
            raise HTTPException(status_code=403, detail="Только сержант/помощник/админ могут создавать опросы")
        if scope_type == "system" and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Системный опрос может создать только админ")
        if scope_type == "group" and user["role"] not in ("sergeant", "admin"):
            raise HTTPException(status_code=403, detail="Опрос по группе может создать только сержант или админ")
        scope_value = "system" if scope_type == "system" else (user["group_name"] if scope_type == "group" else str(user["enrollment_year"]))
        cursor = execute(conn,
            "INSERT INTO custom_surveys (title, scope_type, scope_value, created_by_telegram_id, ends_at) VALUES (?, ?, ?, ?, ?)",
            (title, scope_type, scope_value, telegram_id, ends_at or None)
        )
        survey_id = cursor.lastrowid
        for i, text in enumerate(options):
            if (str(text) or "").strip():
                execute(conn,
                    "INSERT INTO custom_survey_options (survey_id, option_text, sort_order) VALUES (?, ?, ?)",
                    (survey_id, str(text).strip(), i)
                )
        conn.commit()
        conn.close()
        return {"status": "ok", "survey_id": survey_id}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Создание опроса: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.get("/api/survey/custom/{survey_id}")
async def get_custom_survey(survey_id: int, telegram_id: int):
    """Опции опроса, статус завершения, свой голос (если есть)."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        s = execute(conn,
            "SELECT id, title, scope_type, scope_value, created_by_telegram_id, ends_at, completed_at FROM custom_surveys WHERE id = ?",
            (survey_id,)
        ).fetchone()
        if not s:
            raise HTTPException(status_code=404, detail="Опрос не найден")
        options = execute(conn,
            "SELECT id, option_text, sort_order FROM custom_survey_options WHERE survey_id = ? ORDER BY sort_order",
            (survey_id,)
        ).fetchall()
        my_vote = execute(conn,
            "SELECT option_id FROM custom_survey_votes WHERE survey_id = ? AND user_telegram_id = ?",
            (survey_id, telegram_id)
        ).fetchone()
        counts = {}
        for opt in options:
            c = execute(conn,
                "SELECT COUNT(*) FROM custom_survey_votes WHERE survey_id = ? AND option_id = ?",
                (survey_id, opt["id"])
            ).fetchone()
            counts[opt["id"]] = c[0]
        user_row = execute(conn,"SELECT role FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        role = user_row["role"] if user_row else "user"
        can_complete = s["completed_at"] is None and (
            s["created_by_telegram_id"] == telegram_id or role in ("admin", "assistant")
        )
        conn.close()
        return {
            "id": s["id"],
            "title": s["title"],
            "completed_at": s["completed_at"],
            "ends_at": s["ends_at"],
            "created_by_telegram_id": s["created_by_telegram_id"],
            "options": [{"id": o["id"], "text": o["option_text"], "votes": counts.get(o["id"], 0)} for o in options],
            "my_option_id": my_vote["option_id"] if my_vote else None,
            "can_complete": can_complete,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Опрос: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.post("/api/survey/custom/{survey_id}/vote")
async def vote_custom_survey(survey_id: int, data: dict):
    """Проголосовать за один вариант (заменяет предыдущий голос)."""
    telegram_id = data.get("telegram_id")
    option_id = data.get("option_id")
    if not telegram_id or not option_id:
        raise HTTPException(status_code=400, detail="Нужны telegram_id и option_id")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        s = execute(conn,"SELECT id, completed_at FROM custom_surveys WHERE id = ?", (survey_id,)).fetchone()
        if not s:
            raise HTTPException(status_code=404, detail="Опрос не найден")
        if s["completed_at"]:
            raise HTTPException(status_code=400, detail="Опрос уже завершён")
        opt = execute(conn,"SELECT id FROM custom_survey_options WHERE survey_id = ? AND id = ?", (survey_id, option_id)).fetchone()
        if not opt:
            raise HTTPException(status_code=400, detail="Вариант не найден")
        execute(conn,
            "INSERT OR REPLACE INTO custom_survey_votes (survey_id, user_telegram_id, option_id) VALUES (?, ?, ?)",
            (survey_id, telegram_id, option_id)
        )
        conn.commit()
        conn.close()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Голос: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.post("/api/survey/custom/{survey_id}/complete")
async def complete_custom_survey(survey_id: int, data: dict):
    """Завершить опрос досрочно (только создатель или админ/помощник)."""
    telegram_id = data.get("telegram_id")
    if not telegram_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        s = execute(conn,
            "SELECT created_by_telegram_id, completed_at FROM custom_surveys WHERE id = ?", (survey_id,)
        ).fetchone()
        if not s:
            raise HTTPException(status_code=404, detail="Опрос не найден")
        if s["completed_at"]:
            conn.close()
            return {"status": "ok", "message": "Уже завершён"}
        user = execute(conn,"SELECT role FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user or (user["role"] not in ("admin", "assistant") and s["created_by_telegram_id"] != telegram_id):
            raise HTTPException(status_code=403, detail="Завершить может только создатель или админ/помощник")
        from datetime import datetime
        execute(conn,
            "UPDATE custom_surveys SET completed_at = ? WHERE id = ?",
            (datetime.utcnow().isoformat(), survey_id)
        )
        conn.commit()
        conn.close()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Завершение опроса: {e}")
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


# ============================================
# 7. ПРОФИЛЬ, АВАТАР, МАТЕРИАЛЫ
# ============================================

# Папка для аватаров
AVATARS_DIR = os.path.join(os.path.dirname(__file__), "uploads", "avatars")
os.makedirs(AVATARS_DIR, exist_ok=True)

# Папка для материалов занятий
MATERIALS_DIR = os.path.join(os.path.dirname(__file__), "uploads", "materials")
os.makedirs(MATERIALS_DIR, exist_ok=True)


@app.post("/api/profile/avatar")
async def upload_avatar(file: UploadFile = File(...), telegram_id: int = Form(...)):
    """Загрузить аватар пользователя."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Нужен файл изображения")
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        raise HTTPException(status_code=400, detail="Допустимые форматы: jpg, png, webp, gif")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 5МБ)")
    filename = f"{telegram_id}{ext}"
    path = os.path.join(AVATARS_DIR, filename)
    # Remove old avatars for this user
    for old in os.listdir(AVATARS_DIR):
        if old.startswith(str(telegram_id) + "."):
            try:
                os.remove(os.path.join(AVATARS_DIR, old))
            except Exception:
                pass
    with open(path, "wb") as f:
        f.write(contents)
    return {"status": "ok", "avatar_url": f"/uploads/avatars/{filename}"}


@app.get("/api/profile/avatar/{telegram_id}")
async def get_avatar(telegram_id: int):
    """Получить аватар пользователя."""
    for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        path = os.path.join(AVATARS_DIR, f"{telegram_id}{ext}")
        if os.path.isfile(path):
            return FileResponse(path)
    raise HTTPException(status_code=404, detail="Аватар не найден")


@app.get("/api/profile/full")
async def get_full_profile(telegram_id: int):
    """Полная информация для страницы профиля."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        cursor = execute(conn, "PRAGMA table_info(users)")
        cols = [r['name'] for r in cursor.fetchall()]
        name_col = 'fio' if 'fio' in cols else 'full_name'
        group_col = 'group_name' if 'group_name' in cols else None
        
        row = execute(conn,
            f"SELECT telegram_id, {name_col} as fio, enrollment_year, role, status{', ' + group_col + ' as group_name' if group_col else ''} FROM users WHERE telegram_id = ?",
            (telegram_id,)
        ).fetchone()
        if not row:
            return {"error": "Пользователь не найден"}
        
        fio = row['fio'] or ''
        ey = row['enrollment_year']
        group = row['group_name'] if group_col else ''
        role = row.get('role', 'user') or 'user'
        
        try:
            course = get_current_course(int(ey)) if ey else 1
        except Exception:
            course = 1
        
        # Avatar URL
        avatar_url = None
        for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
            if os.path.isfile(os.path.join(AVATARS_DIR, f"{telegram_id}{ext}")):
                avatar_url = f"/uploads/avatars/{telegram_id}{ext}"
                break
        
        # Duty stats
        points = _get_user_duty_points(conn, telegram_id, None, None)
        fio_variants = _fio_match_variants(fio) or [fio or ""]
        placeholders = ",".join(["?"] * len(fio_variants))
        try:
            duty_count = execute(conn,
                f"SELECT COUNT(*) as cnt FROM duty_schedule WHERE fio IN ({placeholders})",
                tuple(fio_variants)
            ).fetchone()
            total_duties = duty_count['cnt'] if duty_count else 0
        except Exception:
            total_duties = 0
        
        # Achievements
        try:
            _unlock_achievements(conn, telegram_id)
            achs = execute(conn, """
                SELECT a.id, a.title, a.description, a.icon_url,
                       ua.telegram_id IS NOT NULL AS unlocked
                FROM achievements a
                LEFT JOIN user_achievements ua ON ua.achievement_id = a.id AND ua.telegram_id = ?
                ORDER BY a.sort_order
            """, (telegram_id,)).fetchall()
            achievements = [{"id": a['id'], "title": a['title'], "description": a['description'],
                           "icon_url": a.get('icon_url', ''), "unlocked": bool(a['unlocked'])} for a in achs]
        except Exception:
            achievements = []
        
        # Sick leave
        sick_leaves = []
        try:
            if 'sick_leave' in [r['name'] for r in execute(conn, "PRAGMA table_info(sick_leave)").fetchall()]:
                pass
        except Exception:
            pass
        try:
            sl_rows = execute(conn,
                "SELECT id, report_date, created_at FROM sick_leave WHERE telegram_id = ? ORDER BY report_date DESC LIMIT 20",
                (telegram_id,)
            ).fetchall()
            sick_leaves = [{"date": r['report_date'], "created_at": r.get('created_at', '')} for r in sl_rows]
        except Exception:
            sick_leaves = []
        
        return {
            "telegram_id": telegram_id,
            "fio": fio,
            "group": group,
            "enrollment_year": ey,
            "course": course,
            "course_label": "Выпускник" if course >= 5 else str(course),
            "role": role,
            "avatar_url": avatar_url,
            "points": points,
            "total_duties": total_duties,
            "achievements": achievements,
            "sick_leaves": sick_leaves,
        }
    finally:
        conn.close()


@app.post("/api/profile/sick-leave")
async def add_sick_leave(data: dict):
    """Добавить/указать больничный с-по."""
    telegram_id = data.get("telegram_id")
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    if not telegram_id or not date_from:
        raise HTTPException(status_code=400, detail="Нужны telegram_id и date_from")
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        if date_to:
            from datetime import date as _date
            d_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
            delta = (d_to - d_from).days
            if delta < 0 or delta > 60:
                raise HTTPException(status_code=400, detail="Некорректный период")
            current = d_from
            while current <= d_to:
                try:
                    execute(conn,
                        "INSERT INTO sick_leave (telegram_id, report_date) VALUES (?, ?)",
                        (telegram_id, current.strftime("%Y-%m-%d"))
                    )
                except Exception:
                    pass
                current += timedelta(days=1)
        else:
            execute(conn,
                "INSERT INTO sick_leave (telegram_id, report_date) VALUES (?, ?)",
                (telegram_id, date_from)
            )
        conn.commit()
        return {"status": "ok"}
    finally:
        conn.close()


@app.get("/api/groups/list")
async def list_groups():
    """Список доступных групп из Апекс-кэша или из БД."""
    groups = {}
    # Try Apex cache first
    cache_path = os.path.join(os.path.dirname(__file__), "apex_cache", "apex_groups.json")
    if os.path.isfile(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                groups = json.load(f)
        except Exception:
            pass
    if not groups:
        conn = get_db()
        if conn:
            try:
                rows = execute(conn,
                    "SELECT DISTINCT group_name FROM users WHERE group_name IS NOT NULL AND group_name != '' ORDER BY group_name"
                ).fetchall()
                groups = {r['group_name']: 0 for r in rows}
            finally:
                conn.close()
    return {"groups": sorted(groups.keys()) if groups else []}


@app.get("/api/schedule/today-by-group")
async def get_today_schedule_by_group(group: str, year: int, date: str = None):
    """Расписание на день по группе (без требования telegram_id / зарегистрированного пользователя)."""
    target_date = datetime.now().date()
    if date:
        try:
            target_date = datetime.strptime(date.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    lessons = []
    message = None
    try:
        parser = _get_apex_parser()
        lessons = parser.get_schedule_for_date(group, year, target_date)
    except HTTPException:
        message = "Сервис расписания временно недоступен"
    except ValueError as ve:
        message = f"Группа не найдена: {ve}"
    except Exception as e:
        print(f"[WARN] schedule by group: {e}")
        message = "Не удалось загрузить расписание"
    return {"date": target_date.strftime("%Y-%m-%d"), "group": group, "year": year, "lessons": lessons, "message": message}


@app.get("/api/schedule/week-by-group")
async def get_week_schedule_by_group(group: str, year: int, date: str = None):
    """Расписание на неделю по группе (без требования telegram_id)."""
    target = datetime.now().date()
    if date:
        try:
            target = datetime.strptime(date.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    weekday = target.weekday()
    monday = target - timedelta(days=weekday)
    week_schedule = {}
    message = None
    try:
        parser = _get_apex_parser()
        week_schedule = parser.get_schedule_for_week(group, year, monday)
    except HTTPException:
        message = "Сервис расписания временно недоступен"
    except ValueError as ve:
        message = f"Группа не найдена: {ve}"
    except Exception as e:
        print(f"[WARN] week schedule by group: {e}")
        message = "Не удалось загрузить расписание"
    return {"week_start": monday.strftime("%Y-%m-%d"), "group": group, "year": year, "schedule": week_schedule, "message": message}


@app.get("/api/rating/top-enhanced")
async def rating_top_enhanced(telegram_id: int, period: str = "all", scope: str = "course", limit: int = 30):
    """Расширенный топ рейтинга с аватарами и изменением за сегодня."""
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="База данных не найдена")
    try:
        user = execute(conn, "SELECT enrollment_year FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        if not user:
            return {"top": []}
        ey = user["enrollment_year"]
        month_from = datetime.now().strftime("%Y-%m") if period == "month" else None
        month_to = datetime.now().strftime("%Y-%m") if period == "month" else None
        if scope == "course":
            rows = execute(conn, "SELECT telegram_id, fio, group_name FROM users WHERE enrollment_year = ? AND status = 'активен'", (ey,)).fetchall()
        else:
            rows = execute(conn, "SELECT telegram_id, fio, group_name FROM users WHERE status = 'активен'").fetchall()
        result = []
        for r in rows:
            tid = r["telegram_id"]
            p = _get_user_duty_points(conn, tid, month_from, month_to)
            avatar_url = None
            for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
                if os.path.isfile(os.path.join(AVATARS_DIR, f"{tid}{ext}")):
                    avatar_url = f"/uploads/avatars/{tid}{ext}"
                    break
            result.append({
                "telegram_id": tid,
                "fio": r["fio"],
                "group_name": r["group_name"],
                "points": p,
                "avatar_url": avatar_url,
            })
        result.sort(key=lambda x: -x["points"])
        result = result[:limit]
        for i, row in enumerate(result):
            row["rank"] = i + 1
        return {"top": result, "period": period, "scope": scope}
    finally:
        conn.close()


# ============================================
# 6. СТАТИКА И ГЛАВНАЯ (исправлено: не подменяем пути)
# ============================================

# Статика для загруженных файлов (аватары, материалы)
_UPLOADS_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(_UPLOADS_DIR, exist_ok=True)
if os.path.isdir(_UPLOADS_DIR):
    app.mount("/uploads", StaticFiles(directory=_UPLOADS_DIR), name="uploads")

# Сайт (главная «Мой день») — открыть по адресу /site/
_SITE_DIR = os.path.join(os.path.dirname(__file__), "site")
if os.path.isdir(_SITE_DIR):
    app.mount("/site", StaticFiles(directory=_SITE_DIR, html=True), name="site")

@app.get("/app", response_class=HTMLResponse)
async def serve_app():
    file_path = os.path.join("app", "index.html")
    if not os.path.exists(file_path):
        return HTMLResponse("<h1>❌ index.html не найден</h1>", 404)

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    # НИКАКОЙ ПОДМЕНЫ — оставляем пути как в исходном HTML
    return HTMLResponse(content=content)


# ============================================
# 8. ЗАПУСК
# ============================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)